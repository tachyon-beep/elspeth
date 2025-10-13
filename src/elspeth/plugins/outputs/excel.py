"""Result sink that materialises experiment outputs into an Excel workbook."""

from __future__ import annotations

import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping

from elspeth.core.interfaces import Artifact, ArtifactDescriptor, ResultSink
from elspeth.core.security import normalize_determinism_level, normalize_security_level
from elspeth.plugins.outputs._sanitize import sanitize_cell

logger = logging.getLogger(__name__)


def _load_workbook_dependencies():  # type: ignore[return-any]
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover - handled during sink initialisation
        raise RuntimeError("ExcelResultSink requires the 'openpyxl' package. Install with 'pip install openpyxl'") from exc
    return Workbook


class ExcelResultSink(ResultSink):
    """Persist experiment payloads into a timestamped Excel workbook."""

    def __init__(
        self,
        *,
        base_path: str | Path,
        workbook_name: str | None = None,
        timestamped: bool = True,
        results_sheet: str = "Results",
        manifest_sheet: str = "Manifest",
        aggregates_sheet: str = "Aggregates",
        include_manifest: bool = True,
        include_aggregates: bool = True,
        on_error: str = "abort",
        sanitize_formulas: bool = True,
        sanitize_guard: str = "'",
    ) -> None:
        self.base_path = Path(base_path)
        self.workbook_name = workbook_name
        self.timestamped = timestamped
        self.results_sheet = results_sheet
        self.manifest_sheet = manifest_sheet
        self.aggregates_sheet = aggregates_sheet
        self.include_manifest = include_manifest
        self.include_aggregates = include_aggregates
        if on_error not in {"abort", "skip"}:
            raise ValueError("on_error must be 'abort' or 'skip'")
        self.on_error = on_error
        if not sanitize_guard:
            sanitize_guard = "'"
        if len(sanitize_guard) != 1:
            raise ValueError("sanitize_guard must be a single character")
        self.sanitize_formulas = sanitize_formulas
        self.sanitize_guard = sanitize_guard
        if not self.sanitize_formulas:
            logger.warning("Excel sink sanitization disabled; outputs may trigger spreadsheet formulas.")
        # Ensure dependency availability early for fast failure when configured incorrectly.
        self._workbook_factory = _load_workbook_dependencies()
        self._last_workbook_path: str | None = None
        self._security_level: str | None = None
        self._determinism_level: str | None = None
        self._sanitization = {
            "enabled": self.sanitize_formulas,
            "guard": self.sanitize_guard,
        }

    # ------------------------------------------------------------------ public API
    def write(self, results: Dict[str, Any], *, metadata: Dict[str, Any] | None = None) -> None:
        metadata = metadata or {}
        timestamp = datetime.now(timezone.utc)
        try:
            path = self._resolve_path(metadata, timestamp)
            path.parent.mkdir(parents=True, exist_ok=True)

            workbook = self._workbook_factory()
            self._populate_results_sheet(workbook, results.get("results", []))

            if self.include_manifest:
                self._populate_manifest_sheet(workbook, results, metadata, timestamp)

            if self.include_aggregates and results.get("aggregates"):
                self._populate_aggregates_sheet(workbook, results.get("aggregates"))

            workbook.save(path)
            self._last_workbook_path = str(path)
            if metadata:
                self._security_level = normalize_security_level(metadata.get("security_level"))
                self._determinism_level = normalize_determinism_level(metadata.get("determinism_level"))
        except Exception as exc:
            if self.on_error == "skip":
                logger.warning("Excel sink failed; skipping workbook creation: %s", exc)
                return
            raise

    # ------------------------------------------------------------------ helpers
    def _sanitize_value(self, value: Any) -> Any:
        if not self.sanitize_formulas:
            return value
        return sanitize_cell(value, guard=self.sanitize_guard)

    def _sanitize_header(self, value: Any) -> Any:
        if not isinstance(value, str):
            return value
        return self._sanitize_value(value)

    def _resolve_path(self, metadata: Mapping[str, Any], timestamp: datetime) -> Path:
        name = self.workbook_name or str(metadata.get("experiment") or metadata.get("name") or "experiment")
        if name.endswith(".xlsx"):
            name = name.removesuffix(".xlsx")
        if self.timestamped:
            name = f"{name}_{timestamp.strftime('%Y%m%dT%H%M%SZ')}"
        return self.base_path / f"{name}.xlsx"

    def _populate_results_sheet(self, workbook, entries: Iterable[Mapping[str, Any]]) -> None:  # type: ignore[no-untyped-def]
        sheet = workbook.active
        sheet.title = self.results_sheet

        flattened = [self._flatten_result(entry) for entry in entries]
        headers: list[str] = []
        if flattened:
            headers = sorted({key for row in flattened for key in row.keys()})
            sheet.append([self._sanitize_header(column) for column in headers])
            for row in flattened:
                sheet.append([self._sanitize_value(row.get(column)) for column in headers])
        else:
            sheet.append([self._sanitize_value("no_results")])

    def _populate_manifest_sheet(
        self,
        workbook,
        results: Mapping[str, Any],
        metadata: Mapping[str, Any],
        timestamp: datetime,
    ) -> None:  # type: ignore[no-untyped-def]
        sheet = workbook.create_sheet(self.manifest_sheet)
        manifest = self._build_manifest(results, metadata, timestamp)
        sheet.append([self._sanitize_header("key"), self._sanitize_header("value")])
        for key, value in manifest.items():
            if isinstance(value, (dict, list)):
                rendered = json.dumps(value, sort_keys=True)
            else:
                rendered = value
            sheet.append([self._sanitize_value(key), self._sanitize_value(rendered)])

    def _populate_aggregates_sheet(self, workbook, aggregates: Mapping[str, Any]) -> None:  # type: ignore[no-untyped-def]
        sheet = workbook.create_sheet(self.aggregates_sheet)
        sheet.append(
            [
                self._sanitize_header("metric"),
                self._sanitize_header("value"),
            ]
        )
        for key, value in aggregates.items():
            if isinstance(value, Mapping):
                rendered = json.dumps(value, sort_keys=True)
            else:
                rendered = value
            sheet.append([self._sanitize_value(key), self._sanitize_value(rendered)])

    @staticmethod
    def _flatten_result(entry: Mapping[str, Any]) -> Dict[str, Any]:
        flat: Dict[str, Any] = {}
        row = entry.get("row")
        if isinstance(row, Mapping):
            for key, value in row.items():
                flat[f"row.{key}"] = value
        for key, value in entry.items():
            if key == "row":
                continue
            if isinstance(value, (dict, list)):
                flat[key] = json.dumps(value, sort_keys=True)
            else:
                flat[key] = value
        return flat

    def _build_manifest(
        self,
        results: Mapping[str, Any],
        metadata: Mapping[str, Any],
        timestamp: datetime,
    ) -> Dict[str, Any]:
        manifest: Dict[str, Any] = {
            "generated_at": timestamp.isoformat(),
            "rows": (len(results.get("results", [])) if isinstance(results.get("results"), list) else 0),
            "metadata": dict(metadata),
            "sanitization": self._sanitization,
        }
        if "cost_summary" in results:
            manifest["cost_summary"] = results["cost_summary"]
        if "failures" in results:
            manifest["failures"] = results["failures"]
        return manifest

    def produces(self):  # pragma: no cover - placeholder for artifact chaining
        return [
            ArtifactDescriptor(name="excel", type="file/xlsx", persist=True, alias="excel"),
        ]

    def consumes(self):  # pragma: no cover - placeholder for artifact chaining
        return []

    def finalize(self, artifacts, *, metadata=None):  # pragma: no cover - optional cleanup
        return None

    def collect_artifacts(self) -> Dict[str, Artifact]:  # pragma: no cover
        if not self._last_workbook_path:
            return {}
        artifact = Artifact(
            id="",
            type="file/xlsx",
            path=self._last_workbook_path,
            metadata={
                "path": self._last_workbook_path,
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "security_level": self._security_level,
                "determinism_level": self._determinism_level,
                "sanitization": self._sanitization,
            },
            persist=True,
            security_level=self._security_level,
            determinism_level=self._determinism_level,
        )
        self._last_workbook_path = None
        self._security_level = None
        self._determinism_level = None
        return {"excel": artifact}
