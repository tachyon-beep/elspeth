"""Result sink that materialises experiment outputs into an Excel workbook."""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook  # type: ignore[import-untyped]

from elspeth.core.base.plugin import BasePlugin
from elspeth.core.base.protocols import Artifact, ArtifactDescriptor, ResultSink
from elspeth.core.base.types import DeterminismLevel, SecurityLevel
from elspeth.core.security import ensure_determinism_level, ensure_security_level
from elspeth.core.utils.path_guard import resolve_under_base, safe_atomic_write
from elspeth.plugins.nodes.sinks._sanitize import sanitize_cell

logger = logging.getLogger(__name__)


@dataclass
class ExcelSinkConfig:
    """Configuration for ExcelResultSink to reduce constructor parameter count."""

    base_path: str | Path
    workbook_name: str | None = None
    timestamped: bool = True
    results_sheet: str = "Results"
    manifest_sheet: str = "Manifest"
    aggregates_sheet: str = "Aggregates"
    include_manifest: bool = True
    include_aggregates: bool = True
    on_error: str = "abort"
    sanitize_formulas: bool = True
    sanitize_guard: str = "'"


def _load_workbook_dependencies() -> Any:
    """Return the Workbook class for creating Excel files.

    Returns:
        Workbook class from openpyxl
    """
    return Workbook


class ExcelResultSink(BasePlugin, ResultSink):
    """Persist experiment payloads into a timestamped Excel workbook.

    Inherits from BasePlugin to provide security enforcement (ADR-004).
    """

    def __init__(
        self,
        *,
        base_path: str | Path,
        workbook_name: str | None = None,
        timestamped: bool = True,
        results_sheet: str = "Results",
        manifest_sheet: str = "Manifest",
        aggregates_sheet: str = "Aggregates",
        include_manifest: bool = True,
        include_aggregates: bool = True,
        on_error: str = "abort",
        sanitize_formulas: bool = True,
        sanitize_guard: str = "'",
        config: ExcelSinkConfig | None = None,
        allowed_base_path: str | Path | None = None,
        security_level: SecurityLevel,  # REQUIRED - no default (ADR-004 requirement)
        allow_downgrade: bool = True,  # ADR-005: Trusted downgrade for sinks (explicit choice, matches default suite)
    ) -> None:
        """Initialize Excel sink.

        Args can be provided either directly or via config object.
        If config is provided, it takes precedence over individual args.
        This supports both legacy and new usage patterns.
        """
        # Initialize BasePlugin with security level and downgrade policy (ADR-004, ADR-005)
        super().__init__(security_level=security_level, allow_downgrade=allow_downgrade)

        # Use config if provided, otherwise use individual args
        if config is not None:
            base_path = config.base_path
            workbook_name = config.workbook_name
            timestamped = config.timestamped
            results_sheet = config.results_sheet
            manifest_sheet = config.manifest_sheet
            aggregates_sheet = config.aggregates_sheet
            include_manifest = config.include_manifest
            include_aggregates = config.include_aggregates
            on_error = config.on_error
            sanitize_formulas = config.sanitize_formulas
            sanitize_guard = config.sanitize_guard
        self.base_path = Path(base_path)
        self.workbook_name = workbook_name
        self.timestamped = timestamped
        self.results_sheet = results_sheet
        self.manifest_sheet = manifest_sheet
        self.aggregates_sheet = aggregates_sheet
        self.include_manifest = include_manifest
        self.include_aggregates = include_aggregates
        if on_error not in {"abort", "skip"}:
            raise ValueError("on_error must be 'abort' or 'skip'")
        self.on_error = on_error
        if not sanitize_guard:
            sanitize_guard = "'"
        if len(sanitize_guard) != 1:
            raise ValueError("sanitize_guard must be a single character")
        self.sanitize_formulas = sanitize_formulas
        self.sanitize_guard = sanitize_guard
        if not self.sanitize_formulas:
            logger.warning("Excel sink sanitization disabled; outputs may trigger spreadsheet formulas.")
        # Ensure dependency availability early for fast failure when configured incorrectly.
        self._workbook_factory = _load_workbook_dependencies()
        self._last_workbook_path: str | None = None
        # Runtime data classification tracking (separate from sink's security clearance)
        # security_level (from BasePlugin) = sink's clearance level
        # _artifact_security_level = runtime classification of written data (for artifact metadata)
        self._artifact_security_level: SecurityLevel | None = None
        self._artifact_determinism_level: DeterminismLevel | None = None
        self._sanitization = {
            "enabled": self.sanitize_formulas,
            "guard": self.sanitize_guard,
        }
        # Allowed base directory for writes; default to ./outputs
        try:
            default_base = Path(base_path).resolve()
            self._allowed_base = Path(allowed_base_path).resolve() if allowed_base_path is not None else default_base
        except Exception:  # pragma: no cover - defensive
            self._allowed_base = Path.cwd().resolve()

    # ------------------------------------------------------------------ public API
    def write(self, results: dict[str, Any], *, metadata: dict[str, Any] | None = None) -> None:
        metadata = metadata or {}
        timestamp = datetime.now(timezone.utc)
        try:
            path = self._resolve_path(metadata, timestamp)
            target = resolve_under_base(path, self._allowed_base)

            plugin_logger = getattr(self, "plugin_logger", None)
            if plugin_logger:
                plugin_logger.log_event(
                    "sink_write_attempt",
                    message=f"Excel write attempt: {target}",
                    metadata={"path": str(target)},
                )

            workbook = self._workbook_factory()
            self._populate_results_sheet(workbook, results.get("results", []))

            if self.include_manifest:
                self._populate_manifest_sheet(workbook, results, metadata, timestamp)

            if self.include_aggregates and results.get("aggregates"):
                aggregates = results.get("aggregates")
                if isinstance(aggregates, Mapping):
                    self._populate_aggregates_sheet(workbook, aggregates)

            def _writer(tmp_path: Path) -> None:
                workbook.save(tmp_path)

            safe_atomic_write(target, _writer)
            self._last_workbook_path = str(target)
            if metadata:
                level = metadata.get("security_level")
                det = metadata.get("determinism_level")
                self._artifact_security_level = level if isinstance(level, SecurityLevel) else ensure_security_level(level)
                self._artifact_determinism_level = det if isinstance(det, DeterminismLevel) else ensure_determinism_level(det)
            if plugin_logger:
                try:
                    size = Path(self._last_workbook_path).stat().st_size if self._last_workbook_path else 0
                except Exception:
                    size = 0
                plugin_logger.log_event(
                    "sink_write",
                    message=f"Excel written to {target}",
                    metrics={"bytes": size},
                    metadata={"path": str(target)},
                )
        except Exception as exc:
            if self.on_error == "skip":
                logger.warning("Excel sink failed; skipping workbook creation: %s", exc)
                plugin_logger = getattr(self, "plugin_logger", None)
                if plugin_logger:
                    plugin_logger.log_error(exc, context="excel sink write", recoverable=True)
                return
            raise

    # ------------------------------------------------------------------ helpers
    def _sanitize_value(self, value: Any) -> Any:
        if not self.sanitize_formulas:
            return value
        return sanitize_cell(value, guard=self.sanitize_guard)

    def _sanitize_header(self, value: Any) -> Any:
        if not isinstance(value, str):
            return value
        return self._sanitize_value(value)

    def _resolve_path(self, metadata: Mapping[str, Any], timestamp: datetime) -> Path:
        name = self.workbook_name or str(metadata.get("experiment") or metadata.get("name") or "experiment")
        if name.endswith(".xlsx"):
            name = name.removesuffix(".xlsx")
        if self.timestamped:
            name = f"{name}_{timestamp.strftime('%Y%m%dT%H%M%SZ')}"
        return self.base_path / f"{name}.xlsx"

    def _populate_results_sheet(self, workbook: Any, entries: Iterable[Mapping[str, Any]]) -> None:
        sheet = workbook.active
        sheet.title = self.results_sheet

        flattened = [self._flatten_result(entry) for entry in entries]
        headers: list[str] = []
        if flattened:
            headers = sorted({key for row in flattened for key in row.keys()})
            sheet.append([self._sanitize_header(column) for column in headers])
            for row in flattened:
                sheet.append([self._sanitize_value(row.get(column)) for column in headers])
        else:
            sheet.append([self._sanitize_value("no_results")])

    def _populate_manifest_sheet(
        self,
        workbook: Any,
        results: Mapping[str, Any],
        metadata: Mapping[str, Any],
        timestamp: datetime,
    ) -> None:
        sheet = workbook.create_sheet(self.manifest_sheet)
        manifest = self._build_manifest(results, metadata, timestamp)
        sheet.append([self._sanitize_header("key"), self._sanitize_header("value")])
        for key, value in manifest.items():
            if isinstance(value, (dict, list)):
                rendered = json.dumps(value, sort_keys=True)
            else:
                rendered = value
            sheet.append([self._sanitize_value(key), self._sanitize_value(rendered)])

    def _populate_aggregates_sheet(self, workbook: Any, aggregates: Mapping[str, Any]) -> None:
        sheet = workbook.create_sheet(self.aggregates_sheet)
        sheet.append(
            [
                self._sanitize_header("metric"),
                self._sanitize_header("value"),
            ]
        )
        for key, value in aggregates.items():
            if isinstance(value, Mapping):
                rendered = json.dumps(value, sort_keys=True)
            else:
                rendered = value
            sheet.append([self._sanitize_value(key), self._sanitize_value(rendered)])

    @staticmethod
    def _flatten_result(entry: Mapping[str, Any]) -> dict[str, Any]:
        flat: dict[str, Any] = {}
        row = entry.get("row")
        if isinstance(row, Mapping):
            for key, value in row.items():
                flat[f"row.{key}"] = value
        for key, value in entry.items():
            if key == "row":
                continue
            if isinstance(value, (dict, list)):
                flat[key] = json.dumps(value, sort_keys=True)
            else:
                flat[key] = value
        return flat

    def _build_manifest(
        self,
        results: Mapping[str, Any],
        metadata: Mapping[str, Any],
        timestamp: datetime,
    ) -> dict[str, Any]:
        manifest: dict[str, Any] = {
            "generated_at": timestamp.isoformat(),
            "rows": (len(results.get("results", [])) if isinstance(results.get("results"), list) else 0),
            "metadata": dict(metadata),
            "sanitization": self._sanitization,
        }
        if "cost_summary" in results:
            manifest["cost_summary"] = results["cost_summary"]
        if "failures" in results:
            manifest["failures"] = results["failures"]
        return manifest

    def produces(self) -> list[ArtifactDescriptor]:  # pragma: no cover - placeholder for artifact chaining
        return [
            ArtifactDescriptor(name="excel", type="file/xlsx", persist=True, alias="excel"),
        ]

    def consumes(self) -> list[str]:  # pragma: no cover - placeholder for artifact chaining
        return []

    def finalize(
        self, artifacts: Mapping[str, Artifact], *, metadata: dict[str, Any] | None = None
    ) -> None:  # pragma: no cover - optional cleanup
        return None

    def collect_artifacts(self) -> dict[str, Artifact]:  # pragma: no cover
        if not self._last_workbook_path:
            return {}
        artifact = Artifact(
            id="",
            type="file/xlsx",
            path=self._last_workbook_path,
            metadata={
                "path": self._last_workbook_path,
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "security_level": self._artifact_security_level,
                "determinism_level": self._artifact_determinism_level,
                "sanitization": self._sanitization,
            },
            persist=True,
            security_level=self._artifact_security_level,
            determinism_level=self._artifact_determinism_level,
        )
        self._last_workbook_path = None
        self._artifact_security_level = None
        self._artifact_determinism_level = None
        return {"excel": artifact}
