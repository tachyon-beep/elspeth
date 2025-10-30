import csv
import os
import sys
from pathlib import Path
from typing import Any, Callable

import pytest

from elspeth.core.base.protocols import ResultSink
from elspeth.plugins.nodes.sinks._sanitize import DANGEROUS_PREFIXES

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


@pytest.fixture(scope="session", autouse=True)
def _configure_sidecar_mode_for_tests():
    """Configure sidecar mode for test environments.

    - If ELSPETH_RUN_INTEGRATION_TESTS=1: Preserve sidecar mode (daemon available in CI)
    - Otherwise: Use standalone mode (no daemon required for unit tests)

    Production containers default to sidecar mode (fail-closed security).
    Test/dev environments can explicitly opt-in to standalone mode when the
    sidecar daemon is not available.

    This fixture runs once per test session before any tests execute.
    """
    # Only override to standalone if NOT running integration tests
    if os.environ.get("ELSPETH_RUN_INTEGRATION_TESTS") != "1":
        os.environ["ELSPETH_SIDECAR_MODE"] = "standalone"
    # else: preserve existing ELSPETH_SIDECAR_MODE from environment (sidecar)
    yield
    # Cleanup not needed - environment persists for entire test session


_SANITIZATION_PREFIXES = tuple(prefix for prefix in DANGEROUS_PREFIXES if prefix != "'")
_BOM = "\ufeff"

try:  # pragma: no cover - optional dependency for Excel checks
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None


@pytest.fixture
def assert_sanitized_artifact() -> Callable[[str | Path], None]:
    """Assert that a CSV/Excel artifact contains no unguarded dangerous prefixes."""

    def _check_value(raw: str, *, context: str) -> None:
        value = raw.lstrip(_BOM)
        if not value:
            return
        if value[0] in _SANITIZATION_PREFIXES:
            raise AssertionError(f"Unsanitized spreadsheet value {raw!r} in {context}")

    def _assert(path: str | Path, *, guard: str = "'") -> None:  # guard kept for future extension
        artifact_path = Path(path)
        suffix = artifact_path.suffix.lower()
        if suffix == ".csv":
            with artifact_path.open(encoding="utf-8", newline="") as handle:
                reader = csv.reader(handle)
                for row_index, row in enumerate(reader):
                    for column_index, cell in enumerate(row):
                        _check_value(
                            cell,
                            context=f"{artifact_path.name}:{row_index}:{column_index}",
                        )
        elif suffix in {".xlsx", ".xlsm"}:
            if load_workbook is None:
                raise AssertionError("openpyxl is required to validate Excel artifacts")
            workbook = load_workbook(artifact_path, data_only=False)
            for sheet in workbook.worksheets:
                for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                    for column_index, cell in enumerate(row):
                        if isinstance(cell, str):
                            _check_value(
                                cell,
                                context=f"{artifact_path.name}:{sheet.title}:{row_index}:{column_index}",
                            )
        else:
            raise AssertionError(f"Unsupported artifact type for sanitization check: {artifact_path}")

    return _assert


class SimpleLLM:
    """Deterministic LLM for testing.

    Returns predictable responses for characterization and safety tests.
    Uses row_id from metadata to create unique per-row responses.
    """

    def generate(
        self,
        *,
        system_prompt: str,
        user_prompt: str,
        metadata: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """Generate deterministic response based on row_id from metadata."""
        metadata = metadata or {}
        row_id = metadata.get("row_id", "unknown")
        return {
            "content": f"response_{row_id}",
            "raw": {"usage": {"prompt_tokens": 10, "completion_tokens": 5}},
        }


@pytest.fixture
def simple_llm() -> SimpleLLM:
    """Fixture providing a deterministic LLM for testing."""
    return SimpleLLM()


class CollectingSink(ResultSink):
    """Sink that records all write() calls for test assertions.

    Used across characterization, integration, and behavioral tests to verify
    sink data flow without actual I/O operations.

    Attributes:
        calls: List of (results_dict, metadata_dict) tuples for each write() call
        _elspeth_security_level: Required security level attribute
    """

    def __init__(self) -> None:
        """Initialize collecting sink with empty call log."""
        self.calls: list[tuple[dict[str, Any], dict[str, Any] | None]] = []
        self._elspeth_security_level = "official"

    def write(self, results: dict[str, Any], *, metadata: dict[str, Any] | None = None) -> None:
        """Record write call with results and metadata."""
        self.calls.append((results, metadata))


@pytest.fixture(scope="session", autouse=True)
def _register_test_sinks():
    """Register CollectingSink as a test plugin for sink resolution tests.

    This enables sink definition-based tests (paths 1-3 in test_suite_runner_sink_resolution.py)
    to test the complete integration path: definition → registry → instantiated object.

    Without this, only pre-instantiated sink tests (paths 4-5) would work, leaving a 60%
    coverage gap in the 5-level sink resolution priority chain.
    """
    from elspeth.core.base.plugin_context import PluginContext
    from elspeth.core.registries.sink import sink_registry

    def _create_collecting_sink(options: dict[str, Any], context: PluginContext) -> CollectingSink:
        """Factory function for CollectingSink plugin."""
        return CollectingSink()

    # Minimal schema - accepts any options for test flexibility
    schema = {
        "type": "object",
        "properties": {},
        "additionalProperties": True,
    }

    sink_registry.register("collecting", _create_collecting_sink, schema=schema, declared_security_level="UNOFFICIAL")

    yield  # Test run happens here

    # No cleanup needed - registry state doesn't persist across test runs


class MiddlewareHookTracer:
    """Captures all middleware lifecycle hook calls for verification.

    This tracer implements all suite-level and experiment-level middleware hooks
    to enable precise testing of hook ordering, deduplication, and argument passing
    in suite_runner.py.

    Critical for verifying:
    - Hook call sequence (suite_loaded → experiment_start → complete → suite_complete)
    - Shared middleware deduplication (on_suite_loaded called once per instance)
    - Correct arguments passed to each hook
    - Hook timing (baseline comparison only after baseline completes)

    Usage:
        tracer = MiddlewareHookTracer()
        # ... run suite with tracer as middleware
        assert tracer.get_call_sequence() == ["on_suite_loaded", "on_experiment_start", ...]
        assert tracer.get_suite_loaded_count() == 1  # Deduplication verified
    """

    def __init__(self, name: str = "tracer") -> None:
        """Initialize tracer with optional name for multi-tracer scenarios."""
        self.name = name
        self.calls: list[dict[str, Any]] = []

    def on_suite_loaded(
        self,
        suite_metadata: list[dict[str, Any]],
        preflight_info: dict[str, Any],
    ) -> None:
        """Hook: Suite execution started (should be called once per unique instance)."""
        self.calls.append({
            "hook": "on_suite_loaded",
            "instance_id": id(self),
            "instance_name": self.name,
            "suite_metadata": suite_metadata,
            "preflight_info": preflight_info,
            "experiment_count": len(suite_metadata),
        })

    def on_experiment_start(
        self,
        experiment_name: str,
        metadata: dict[str, Any],
    ) -> None:
        """Hook: Experiment execution started."""
        self.calls.append({
            "hook": "on_experiment_start",
            "instance_id": id(self),
            "instance_name": self.name,
            "experiment": experiment_name,
            "metadata": metadata,
        })

    def on_experiment_complete(
        self,
        experiment_name: str,
        payload: dict[str, Any],
        metadata: dict[str, Any],
    ) -> None:
        """Hook: Experiment execution completed."""
        self.calls.append({
            "hook": "on_experiment_complete",
            "instance_id": id(self),
            "instance_name": self.name,
            "experiment": experiment_name,
            "metadata": metadata,
            "has_payload": payload is not None,
            "result_count": len(payload.get("results", [])) if payload else 0,
        })

    def on_baseline_comparison(
        self,
        experiment_name: str,
        comparisons: dict[str, Any],
    ) -> None:
        """Hook: Baseline comparison executed."""
        self.calls.append({
            "hook": "on_baseline_comparison",
            "instance_id": id(self),
            "instance_name": self.name,
            "experiment": experiment_name,
            "comparison_count": len(comparisons),
            "comparison_plugins": list(comparisons.keys()),
        })

    def on_suite_complete(self) -> None:
        """Hook: Suite execution completed."""
        self.calls.append({
            "hook": "on_suite_complete",
            "instance_id": id(self),
            "instance_name": self.name,
        })

    def get_call_sequence(self) -> list[str]:
        """Return hook names in call order.

        Example:
            ["on_suite_loaded", "on_experiment_start", "on_experiment_complete", ...]
        """
        return [call["hook"] for call in self.calls]

    def get_suite_loaded_count(self) -> int:
        """Count on_suite_loaded calls (should be 1 for shared instances).

        Critical for verifying deduplication: a middleware instance shared across
        multiple experiments should only receive on_suite_loaded ONCE.
        """
        return len([c for c in self.calls if c["hook"] == "on_suite_loaded"])

    def get_experiment_start_count(self) -> int:
        """Count on_experiment_start calls (should equal number of experiments)."""
        return len([c for c in self.calls if c["hook"] == "on_experiment_start"])

    def get_experiment_complete_count(self) -> int:
        """Count on_experiment_complete calls (should equal number of experiments)."""
        return len([c for c in self.calls if c["hook"] == "on_experiment_complete"])

    def get_baseline_comparison_count(self) -> int:
        """Count on_baseline_comparison calls (should equal non-baseline experiments)."""
        return len([c for c in self.calls if c["hook"] == "on_baseline_comparison"])

    def get_suite_complete_count(self) -> int:
        """Count on_suite_complete calls (should be 1)."""
        return len([c for c in self.calls if c["hook"] == "on_suite_complete"])

    def get_experiments_for_hook(self, hook_name: str) -> list[str]:
        """Get experiment names for a specific hook type.

        Args:
            hook_name: Hook to query (e.g., "on_experiment_start")

        Returns:
            List of experiment names in call order
        """
        return [
            call["experiment"]
            for call in self.calls
            if call["hook"] == hook_name and "experiment" in call
        ]

    def verify_hook_ordering(self, expected_sequence: list[str]) -> bool:
        """Verify hooks were called in expected order.

        Args:
            expected_sequence: List of hook names in expected order

        Returns:
            True if sequence matches exactly
        """
        return self.get_call_sequence() == expected_sequence

    def reset(self) -> None:
        """Clear all recorded calls (useful for multi-test scenarios)."""
        self.calls = []


@pytest.fixture
def middleware_tracer() -> MiddlewareHookTracer:
    """Fixture providing a middleware hook tracer for suite_runner testing."""
    return MiddlewareHookTracer()
