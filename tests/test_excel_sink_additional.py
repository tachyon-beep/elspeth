from __future__ import annotations

from pathlib import Path

import openpyxl  # type: ignore[import-untyped]
import pytest

from elspeth.plugins.nodes.sinks.excel import ExcelResultSink


def _results_with_aggregates() -> dict:
    return {
        "results": [{"row": {"a": 1}, "response": {"content": "ok"}}],
        "aggregates": {"score_stats": {"criteria": {"x": {"mean": 0.5}}}},
    }


def test_excel_includes_manifest_and_aggregates(tmp_path: Path) -> None:
    sink = ExcelResultSink(base_path=tmp_path, workbook_name="wb", timestamped=False)
    sink._allowed_base = tmp_path.resolve()  # type: ignore[attr-defined]
    sink.write(_results_with_aggregates(), metadata={"experiment": "e"})
    path = tmp_path / "wb.xlsx"
    assert path.exists()
    wb = openpyxl.load_workbook(path)
    assert "Manifest" in wb.sheetnames
    assert "Aggregates" in wb.sheetnames


def test_excel_on_error_skip_when_save_fails(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    sink = ExcelResultSink(base_path=tmp_path, workbook_name="bad", timestamped=False, on_error="skip")
    sink._allowed_base = tmp_path.resolve()  # type: ignore[attr-defined]

    class _WB:
        def save(self, _path):
            raise RuntimeError("cannot save")

    # Patch workbook factory to produce failing workbook
    monkeypatch.setattr(sink, "_workbook_factory", _WB)
    sink.write({"results": []}, metadata={"experiment": "e"})
    assert not (tmp_path / "bad.xlsx").exists()
