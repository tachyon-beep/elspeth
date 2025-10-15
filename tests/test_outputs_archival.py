import csv
import json
from pathlib import Path
from zipfile import ZipFile

import pytest

from elspeth.plugins.nodes.sinks.excel import ExcelResultSink
from elspeth.plugins.nodes.sinks.zip_bundle import ZipResultSink


def sample_results():
    return {
        "results": [
            {
                "row": {"APPID": "1", "name": "alpha", "formula": "=SUM(1,2)"},
                "response": {"content": "ok"},
                "metrics": {"latency": 0.5},
            }
        ],
        "aggregates": {"score": {"mean": 0.7}},
        "cost_summary": {"total": 1.23},
    }


def test_excel_result_sink_writes_workbook(tmp_path, assert_sanitized_artifact):
    from openpyxl import load_workbook

    sink = ExcelResultSink(
        base_path=tmp_path,
        workbook_name="report",
        include_manifest=True,
        include_aggregates=True,
    )

    sink.write(sample_results(), metadata={"experiment": "exp1"})

    files = list(Path(tmp_path).glob("report_*.xlsx"))
    assert files, "Workbook was not created"
    workbook_path = files[0]
    wb = load_workbook(workbook_path)

    assert sink.results_sheet in wb.sheetnames
    assert sink.manifest_sheet in wb.sheetnames
    assert sink.aggregates_sheet in wb.sheetnames

    ws = wb[sink.results_sheet]
    headers = [cell.value for cell in ws[1]]
    assert "row.APPID" in headers
    assert "row.formula" in headers
    content_cell = ws[2][headers.index("response")]
    assert "ok" in content_cell.value
    formula_cell = ws[2][headers.index("row.formula")]
    assert formula_cell.value == "'=SUM(1,2)"

    manifest_sheet = wb[sink.manifest_sheet]
    manifest_values = {row[0].value: row[1].value for row in manifest_sheet.iter_rows(min_row=2)}
    assert manifest_values["rows"] == 1
    assert json.loads(manifest_values["metadata"])["experiment"] == "exp1"
    sanitization_meta = json.loads(manifest_values["sanitization"])
    assert sanitization_meta == {"enabled": True, "guard": "'"}

    assert_sanitized_artifact(workbook_path)


def test_excel_sink_skip_on_error(monkeypatch, tmp_path, caplog):
    sink = ExcelResultSink(base_path=tmp_path, on_error="skip")

    class DummySheet:
        def __init__(self):
            self.title = ""

        def append(self, values):
            pass

    class DummyWorkbook:
        def __init__(self):
            self.active = DummySheet()

        def create_sheet(self, name):
            sheet = DummySheet()
            sheet.title = name
            return sheet

        def save(self, path):
            raise RuntimeError("boom")

    sink._workbook_factory = lambda: DummyWorkbook()  # type: ignore[assignment]

    with caplog.at_level("WARNING"):
        sink.write(sample_results(), metadata={"experiment": "exp1"})

    assert "skipping workbook" in "".join(caplog.messages)


def test_zip_result_sink_creates_archive(tmp_path):
    sink = ZipResultSink(
        base_path=tmp_path,
        bundle_name="bundle",
        include_csv=True,
    )

    sink.write(sample_results(), metadata={"experiment": "exp1"})

    archives = list(Path(tmp_path).glob("bundle_*.zip"))
    assert archives, "ZIP archive was not created"
    archive_path = archives[0]

    with ZipFile(archive_path) as zf:
        names = set(zf.namelist())
        assert sink.results_name in names
        assert sink.manifest_name in names
        assert sink.csv_name in names

        manifest = json.loads(zf.read(sink.manifest_name))
        assert manifest["rows"] == 1
        assert manifest["metadata"]["experiment"] == "exp1"
        assert manifest["sanitization"] == {"enabled": True, "guard": "'"}

        csv_reader = csv.reader(zf.read(sink.csv_name).decode("utf-8").splitlines())
        header = next(csv_reader)
        rows = list(csv_reader)
        assert header[2] == "formula"
        assert rows
        formula_values = [row[2] for row in rows]
        assert all(value == "'=SUM(1,2)" for value in formula_values)


def test_zip_sink_disable_sanitization(tmp_path):
    sink = ZipResultSink(
        base_path=tmp_path,
        bundle_name="bundle_no_guard",
        include_csv=True,
        sanitize_formulas=False,
        timestamped=False,
    )

    sink.write(sample_results(), metadata={"experiment": "exp1"})

    archive_path = tmp_path / "bundle_no_guard.zip"
    assert archive_path.exists()

    with ZipFile(archive_path) as zf:
        manifest = json.loads(zf.read(sink.manifest_name))
        assert manifest["sanitization"] == {"enabled": False, "guard": "'"}
        csv_reader = csv.reader(zf.read(sink.csv_name).decode("utf-8").splitlines())
        header = next(csv_reader)
        rows = list(csv_reader)
        formula_values = [row[header.index("formula")] for row in rows]
        assert any(value.startswith("=") for value in formula_values)


def test_zip_sink_skip_on_error(monkeypatch, tmp_path, caplog):
    sink = ZipResultSink(base_path=tmp_path, on_error="skip")

    class BrokenZip:
        def __init__(self, *args, **kwargs):
            raise RuntimeError("boom")

    monkeypatch.setattr("elspeth.plugins.nodes.sinks.zip_bundle.ZipFile", BrokenZip)

    with caplog.at_level("WARNING"):
        sink.write(sample_results(), metadata={"experiment": "exp1"})

    assert "skipping archive" in "".join(caplog.messages)
