#!/usr/bin/env python3
"""Compatibility matrix runner for spreadsheet sanitisation outputs."""

from __future__ import annotations

import argparse
import csv as csv_module
import json
import sys
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, Dict, List, Tuple

import pandas as pd

from elspeth.plugins.outputs._sanitize import sanitize_cell
from elspeth.plugins.outputs.csv_file import CsvResultSink
from elspeth.plugins.outputs.excel import ExcelResultSink

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None


DANGEROUS_HEADER = "=danger_header"


def _normalize(value: str) -> str:
    """Normalize newlines for comparisons."""

    return value.replace("\r", "\n")


def _build_results() -> Tuple[Dict[str, Any], List[str]]:
    dangerous_values = [
        "=SUM(A1:A2)",
        "+2",
        "-3",
        "@cmd",
        "\tTabbed",
        "\rCarriage",
        "\nNewline",
        "'AlreadyGuarded",
        "Ordinary text",
        "12345",
    ]
    entries: List[Dict[str, Any]] = []
    for idx, value in enumerate(dangerous_values, start=1):
        entries.append(
            {
                "row": {
                    "id": idx,
                    "input": value,
                    DANGEROUS_HEADER: value,
                },
                "response": {"content": value},
                "responses": {"alt": {"content": value}},
            }
        )
    return {"results": entries, "aggregates": {"count": len(entries)}}, dangerous_values


def _verify_csv(csv_path: Path, guard: str, originals: List[str]) -> List[str]:
    failures: List[str] = []
    df = pd.read_csv(csv_path, dtype=str).fillna("")

    sanitized_header = sanitize_cell(DANGEROUS_HEADER, guard=guard)
    expected_columns = [
        "id",
        "input",
        sanitized_header,
        "llm_content",
        "llm_alt",
    ]

    if list(df.columns) != expected_columns:
        failures.append(f"CSV columns mismatch: expected {expected_columns}, got {list(df.columns)}")

    for idx, original in enumerate(originals):
        expected_value = sanitize_cell(original, guard=guard)
        row = df.iloc[idx]
        if row["id"] != str(idx + 1):
            failures.append(f"CSV row {idx} id mismatch: expected {idx + 1}, got {row['id']}")
        for column in expected_columns[1:]:
            observed = row[column]
            if _normalize(str(observed)) != _normalize(expected_value):
                failures.append(f"CSV row {idx} column '{column}' expected {expected_value!r}, got {observed!r}")

    with open(csv_path, newline="", encoding="utf-8") as handle:
        reader = csv_module.reader(handle)
        header = next(reader)
        if header != expected_columns:
            failures.append(f"CSV header mismatch via csv.reader: expected {expected_columns}, got {header}")
        for row_index, row in enumerate(reader, start=1):
            expected_value = sanitize_cell(originals[row_index - 1], guard=guard)
            if row[0] != str(row_index):
                failures.append(f"CSV reader row {row_index} id mismatch: expected {row_index}, got {row[0]}")
            for column_index, cell in enumerate(row[1:], start=1):
                if _normalize(cell) != _normalize(expected_value):
                    failures.append(f"CSV reader row {row_index} column {column_index} expected {expected_value!r}, got {cell!r}")

    return failures


def _verify_excel(xlsx_path: Path, guard: str, originals: List[str]) -> List[str]:
    if load_workbook is None:
        return ["openpyxl not installed; Excel validation skipped"]

    failures: List[str] = []
    workbook = load_workbook(xlsx_path, data_only=False)
    results_sheet = workbook["Results"]
    header_row = [cell for cell in next(results_sheet.iter_rows(values_only=True))]
    header_index = {name: idx for idx, name in enumerate(header_row)}
    required_columns = {"row.id", "row.input", f"row.{DANGEROUS_HEADER}"}
    missing = sorted(required_columns - header_index.keys())
    if missing:
        failures.append(f"Excel missing expected columns: {missing}")
        return failures

    for row_index, row in enumerate(results_sheet.iter_rows(min_row=2, values_only=True), start=1):
        expected_value = sanitize_cell(originals[row_index - 1], guard=guard)
        row_cells = list(row)

        id_cell = row_cells[header_index["row.id"]]
        if str(id_cell) != str(row_index):
            failures.append(f"Excel row {row_index} id mismatch: expected {row_index}, got {id_cell}")

        input_cell = row_cells[header_index["row.input"]]
        if _normalize(str(input_cell)) != _normalize(expected_value):
            failures.append(f"Excel row {row_index} input mismatch: expected {expected_value!r}, got {input_cell!r}")

        danger_cell = row_cells[header_index[f"row.{DANGEROUS_HEADER}"]]
        if _normalize(str(danger_cell)) != _normalize(expected_value):
            failures.append(f"Excel row {row_index} dangerous header mismatch: expected {expected_value!r}, got {danger_cell!r}")

    if "Manifest" in workbook.sheetnames:
        manifest_sheet = workbook["Manifest"]
        manifest: Dict[str, Any] = {}
        for key, value in manifest_sheet.iter_rows(min_row=2, values_only=True):
            manifest[key] = value
        sanitization_entry = manifest.get("sanitization")
        if isinstance(sanitization_entry, str):
            try:
                sanitization_entry = json.loads(sanitization_entry)
            except json.JSONDecodeError as exc:  # pragma: no cover - defensive
                failures.append(f"Manifest sanitization entry malformed JSON: {exc}")
                sanitization_entry = None
        if not isinstance(sanitization_entry, dict):
            failures.append("Manifest missing sanitization metadata")
        else:
            if not sanitization_entry.get("enabled", False):
                failures.append("Manifest reports sanitization disabled")
            if sanitization_entry.get("guard") != guard:
                failures.append(f"Manifest guard mismatch: expected {guard!r}, got {sanitization_entry.get('guard')!r}")

    return failures


def run_matrix(guard: str, verbose: bool = False) -> int:
    results, originals = _build_results()
    metadata = {"experiment": "compat_matrix"}
    failures: List[str] = []

    with TemporaryDirectory() as tmp:
        tmpdir = Path(tmp)
        csv_path = tmpdir / "results.csv"
        xlsx_path = tmpdir / "results.xlsx"

        csv_sink = CsvResultSink(path=csv_path, sanitize_guard=guard)
        csv_sink.write(results, metadata=metadata)
        failures.extend(_verify_csv(csv_path, guard, originals))

        excel_sink = ExcelResultSink(
            base_path=tmpdir,
            workbook_name="results",
            timestamped=False,
            sanitize_guard=guard,
        )
        excel_sink.write(results, metadata=metadata)
        failures.extend(_verify_excel(xlsx_path, guard, originals))

    reportable = [failure for failure in failures if not failure.startswith("openpyxl")]

    if verbose or reportable:
        for failure in failures:
            print(f"[compat:failure] {failure}")

    return len(reportable)


def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run sanitisation compatibility checks")
    parser.add_argument(
        "--guard",
        default="'",
        help="Guard character to assert (default: apostrophe)",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print detailed failure diagnostics",
    )
    return parser.parse_args(argv)


def main(argv: List[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    if len(args.guard) != 1:
        print("guard must be a single character", file=sys.stderr)
        return 2
    code = run_matrix(args.guard, verbose=args.verbose)
    if code:
        print(f"sanitisation compatibility FAILED ({code} issues)", file=sys.stderr)
        return 1
    print("sanitisation compatibility PASSED")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
