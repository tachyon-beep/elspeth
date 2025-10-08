import numpy as np
import pandas as pd
from scipy import stats
from typing import Dict, List, Tuple, Any, Optional, Union
import json
from collections import Counter
from datetime import datetime
import os
import logging
from dataclasses import dataclass, asdict
from matplotlib import pyplot as plt 

logger = logging.getLogger(__name__)

# Check for optional dependencies at module level
HAS_SKLEARN = False
try:
    from sklearn.metrics import cohen_kappa_score, confusion_matrix
    HAS_SKLEARN = True
except ImportError:
    pass

HAS_PINGOUIN = False
try:
    import pingouin as pg
    HAS_PINGOUIN = True
except ImportError:
    pass

class NumpyJSONEncoder(json.JSONEncoder):
    """JSON encoder that handles NumPy types"""
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, (np.bool_, bool)):
            return bool(obj)
        return super().default(obj)

@dataclass
class StatisticalResult:
    """Structured result for statistical tests"""
    test_name: str
    statistic: float
    p_value: float
    significant: bool
    effect_size: Optional[float] = None
    interpretation: Optional[str] = None

class StatsAnalyzer:
    """Comprehensive statistical analysis across experiments"""
    
    # Criteria name mapping
    CRITERIA_NAMES = {
        0: "circumstance",
        1: "action",
        2: "result", 
        3: "ethical_behaviour",
        4: "sustainability"
    }
    
    # Statistical thresholds
    SIGNIFICANCE_LEVEL = 0.05
    MIN_SAMPLES = 10
    MIN_PAIRED_SAMPLES = 10
    MAX_CACHE_SIZE = 100  # Limit cache size
    
    def __init__(self, all_results: Dict[str, Dict]):
        self.all_results = all_results
        self.baseline_name = self._identify_baseline()
        
        if self.baseline_name and self.baseline_name in all_results:
            self.baseline_data = all_results[self.baseline_name]["results"]
        else:
            logger.warning("No baseline identified, using first experiment")
            self.baseline_name = sorted(all_results.keys())[0] if all_results else None
            self.baseline_data = all_results[self.baseline_name]["results"] if self.baseline_name else []
        
        # Cache for expensive calculations with size management
        self._cache = {}
        self._cache_order = []
    
    def _add_to_cache(self, key: str, value: Any):
        """Add to cache with size limit"""
        if key in self._cache:
            return
        
        if len(self._cache) >= self.MAX_CACHE_SIZE:
            # Remove oldest entry
            oldest_key = self._cache_order.pop(0)
            del self._cache[oldest_key]
        
        self._cache[key] = value
        self._cache_order.append(key)
    
    def _identify_baseline(self) -> Optional[str]:
        """Find the baseline experiment with proper error handling"""
        for name, data in self.all_results.items():
            if "config" in data and hasattr(data.get("config"), "is_baseline"):
                if data["config"].is_baseline:
                    return name
        
        # Default to first experiment or one with "baseline" in name
        for name in sorted(self.all_results.keys()):
            if "baseline" in name.lower():
                return name
        
        return sorted(self.all_results.keys())[0] if self.all_results else None

    def _generate_recommendation(self, results: List[Dict]) -> str:
        """Generate recommendation text based on analysis results"""
        if not results:
            return "Insufficient data for recommendations"
        
        best = results[0]
        if best['significant_after_correction']:
            if best['effect_size'] > 0.5:
                return f"STRONG RECOMMENDATION: Use {best['variant']} - significant improvement with large effect size"
            else:
                return f"RECOMMENDATION: Consider {best['variant']} - statistically significant improvement"
        elif best['statistical_power'] < 0.8:
            needed_n = self.required_sample_size(best['effect_size'])
            return f"INSUFFICIENT DATA: Need ~{needed_n} samples per group for reliable conclusion (currently have ~{len(self.extract_scores(self.baseline_data))})"
        else:
            return f"NO CLEAR WINNER: {best['variant']} shows promise but not statistically significant"

    def _interpret_bayesian(self, prob_improvement: float) -> str:
        """Interpret Bayesian probability of improvement"""
        if prob_improvement > 0.95:
            return "very strong evidence variant is better"
        elif prob_improvement > 0.85:
            return "strong evidence variant is better"
        elif prob_improvement > 0.75:
            return "moderate evidence variant is better"
        elif prob_improvement > 0.65:
            return "weak evidence variant is better"
        elif prob_improvement > 0.35:
            return "no clear difference"
        elif prob_improvement > 0.25:
            return "weak evidence baseline is better"
        elif prob_improvement > 0.15:
            return "moderate evidence baseline is better"
        else:
            return "strong evidence baseline is better"

    def _simple_bayesian(self, baseline_scores: List[int], variant_scores: List[int]) -> Dict:
        """Simplified Bayesian analysis without pymc3"""
        # Use beta-binomial conjugate prior
        baseline_mean = np.mean(baseline_scores)
        variant_mean = np.mean(variant_scores)
        
        # Bootstrap to estimate probability of improvement
        n_bootstrap = 10000
        variant_better = 0
        
        for _ in range(n_bootstrap):
            baseline_sample = np.random.choice(baseline_scores, len(baseline_scores), replace=True)
            variant_sample = np.random.choice(variant_scores, len(variant_scores), replace=True)
            if np.mean(variant_sample) > np.mean(baseline_sample):
                variant_better += 1
        
        prob_improvement = variant_better / n_bootstrap
        
        return {
            'method': 'bootstrap_bayesian',
            'prob_variant_better': prob_improvement,
            'mean_difference': variant_mean - baseline_mean,
            'interpretation': self._interpret_bayesian(prob_improvement)
        }

    def bayesian_comparison(self, baseline_scores: List[int], variant_scores: List[int]) -> Dict:
        """Bayesian analysis - better for small samples"""
        try:
            import pymc3 as pm
            
            with pm.Model() as model:
                # Priors
                mu_baseline = pm.Normal('mu_baseline', mu=3, sd=2)
                mu_variant = pm.Normal('mu_variant', mu=3, sd=2)
                sigma = pm.HalfNormal('sigma', sd=2)
                
                # Likelihoods
                baseline_obs = pm.Normal('baseline_obs', mu=mu_baseline, sd=sigma, observed=baseline_scores)
                variant_obs = pm.Normal('variant_obs', mu=mu_variant, sd=sigma, observed=variant_scores)
                
                # Difference
                diff = pm.Deterministic('difference', mu_variant - mu_baseline)
                
                # Sample
                trace = pm.sample(2000, tune=1000, return_inferencedata=False)
            
            # Calculate probability of improvement
            prob_improvement = (trace['difference'] > 0).mean()
            
            return {
                'method': 'bayesian',
                'prob_variant_better': float(prob_improvement),
                'mean_difference': float(trace['difference'].mean()),
                'ci_95': [float(np.percentile(trace['difference'], 2.5)), 
                        float(np.percentile(trace['difference'], 97.5))],
                'interpretation': self._interpret_bayesian(prob_improvement)
            }
        except ImportError:
            # Fallback to simple Bayesian estimation
            return self._simple_bayesian(baseline_scores, variant_scores)

    def calculate_cliffs_delta(self, group1: List[float], group2: List[float]) -> Tuple[float, str]:
        """Calculate Cliff's Delta - non-parametric effect size for ordinal data"""
        n1, n2 = len(group1), len(group2)
        if n1 == 0 or n2 == 0:
            return 0.0, "no data"
        
        # Count how often values in group2 are higher than in group1
        dominance = sum(1 if y > x else -1 if x > y else 0 
                    for x in group1 for y in group2)
        
        delta = dominance / (n1 * n2)
        
        # Interpretation
        abs_delta = abs(delta)
        if abs_delta < 0.147:
            interpretation = "negligible"
        elif abs_delta < 0.33:
            interpretation = "small"
        elif abs_delta < 0.474:
            interpretation = "medium"
        else:
            interpretation = "large"
        
        return delta, interpretation

    def calculate_krippendorff_alpha(self, baseline_scores: List[int], variant_scores: List[int]) -> float:
        """Calculate Krippendorff's alpha - handles ordinal data and missing values"""
        try:
            import krippendorff
            
            # Format data for Krippendorff
            data = [baseline_scores, variant_scores]
            alpha = krippendorff.alpha(reliability_data=data, level_of_measurement='ordinal')
            return alpha
        except ImportError:
            # Fallback to weighted kappa for ordinal data
            if HAS_SKLEARN:
                from sklearn.metrics import cohen_kappa_score
                return cohen_kappa_score(baseline_scores, variant_scores, weights='quadratic')
            return np.nan

    def analyze_practical_significance(self, variant_name: str) -> Dict:
        """Analyze practical (not just statistical) significance"""
        dist = self.calculate_distribution_shift(variant_name)
        
        if "error" in dist:
            return {"error": dist["error"]}
        
        baseline_scores = self.extract_scores(self.baseline_data)
        variant_scores = self.extract_scores(self.all_results[variant_name]["results"])
        
        # Calculate proportion of meaningful changes (2+ point difference)
        paired = self._get_paired_scores(variant_name)
        meaningful_changes = sum(1 for b, v in paired if abs(b - v) >= 2)
        
        # Calculate NNT (Number Needed to Treat) for improvement
        baseline_success = sum(1 for s in baseline_scores if s >= 4) / len(baseline_scores)
        variant_success = sum(1 for s in variant_scores if s >= 4) / len(variant_scores)
        
        if variant_success > baseline_success:
            nnt = 1 / (variant_success - baseline_success)
        elif variant_success == baseline_success:
            nnt = float('inf')
        else:
            nnt = float('inf')
        
        return {
            'meaningful_change_rate': meaningful_changes / len(paired) if paired else 0,
            'nnt_for_success': nnt,
            'practical_improvement': variant_success - baseline_success,
            'is_practically_significant': (
                (abs(dist['mean_delta']) > 0.5 or 
                meaningful_changes / len(paired) > 0.2) if len(paired) > 0 else False
            )
        }

    def apply_bonferroni_correction(self, p_values: List[float]) -> List[float]:
        """Apply Bonferroni correction for multiple comparisons"""
        n_tests = len(p_values)
        return [min(p * n_tests, 1.0) for p in p_values]

    def apply_fdr_correction(self, p_values: List[float]) -> Tuple[List[float], List[bool]]:
        """Apply False Discovery Rate correction"""
        from statsmodels.stats.multitest import fdrcorrection
        return fdrcorrection(p_values, alpha=self.SIGNIFICANCE_LEVEL)

    def calculate_cohens_d_ci(self, group1: List[float], group2: List[float], 
                            confidence: float = 0.95) -> Tuple[float, float, float]:
        """Calculate Cohen's d with confidence interval using bootstrap"""
        d = self._calculate_cohens_d(group1, group2)
        
        # Bootstrap confidence interval
        n_bootstrap = 1000
        bootstrap_ds = []
        
        for _ in range(n_bootstrap):
            sample1 = np.random.choice(group1, len(group1), replace=True)
            sample2 = np.random.choice(group2, len(group2), replace=True)
            bootstrap_ds.append(self._calculate_cohens_d(sample1, sample2))
        
        alpha = 1 - confidence
        lower = np.percentile(bootstrap_ds, alpha/2 * 100)
        upper = np.percentile(bootstrap_ds, (1 - alpha/2) * 100)
        
        return d, lower, upper

    def ordinal_logistic_regression(self, baseline_scores: List[int], 
                                    variant_scores: List[int]) -> Dict:
        """Perform ordinal logistic regression for proper ordinal data analysis"""
        try:
            from statsmodels.miscmodels.ordinal_model import OrderedModel
            
            # Prepare data
            all_scores = baseline_scores + variant_scores
            groups = [0] * len(baseline_scores) + [1] * len(variant_scores)
            
            df = pd.DataFrame({'score': all_scores, 'group': groups})
            
            # Fit ordinal model
            mod = OrderedModel(df['score'], df[['group']], distr='logit')
            res = mod.fit(method='bfgs')
            
            return {
                'coefficient': res.params[0],
                'p_value': res.pvalues[0],
                'odds_ratio': np.exp(res.params[0]),
                'ci_lower': np.exp(res.conf_int()[0][0]),
                'ci_upper': np.exp(res.conf_int()[0][1])
            }
        except ImportError:
            return {"error": "statsmodels not available for ordinal regression"}

    def calculate_statistical_power(self, effect_size: float, n1: int, n2: int) -> float:
        """Calculate statistical power for detecting the observed effect size"""
        try:
            from statsmodels.stats.power import TTestIndPower
            power_analysis = TTestIndPower()
            power = power_analysis.solve_power(
                effect_size=effect_size,
                nobs1=n1,
                ratio=n2/n1,
                alpha=self.SIGNIFICANCE_LEVEL
            )
            return power
        except:
            # Approximate power calculation
            from scipy.stats import norm
            z_alpha = norm.ppf(1 - self.SIGNIFICANCE_LEVEL/2)
            ncp = effect_size * np.sqrt(n1 * n2 / (n1 + n2))
            power = 1 - norm.cdf(z_alpha - ncp)
            return power

    def required_sample_size(self, effect_size: float, power: float = 0.8) -> int:
        """Calculate required sample size to achieve desired power"""

        if abs(effect_size) < 1e-10:  # Use smaller epsilon
            return 10000

        try:
            from statsmodels.stats.power import TTestIndPower
            power_analysis = TTestIndPower()
            n = power_analysis.solve_power(
                effect_size=effect_size,
                power=power,
                alpha=self.SIGNIFICANCE_LEVEL,
                ratio=1.0
            )
            return int(np.ceil(n))
        except:
            # Approximate calculation
            from scipy.stats import norm
            z_alpha = norm.ppf(1 - self.SIGNIFICANCE_LEVEL/2)
            z_beta = norm.ppf(power)
            n = 2 * ((z_alpha + z_beta) / effect_size) ** 2
            return int(np.ceil(n))

    def test_assumptions(self, scores1: List[float], scores2: List[float]) -> Dict:
        """Test statistical assumptions before running parametric tests"""
        results = {}
        
        # Normality tests
        if len(scores1) >= 3:
            shapiro1 = stats.shapiro(scores1)
            results['group1_normality'] = {
                'statistic': shapiro1.statistic,
                'p_value': shapiro1.pvalue,
                'is_normal': shapiro1.pvalue > 0.05
            }
        
        if len(scores2) >= 3:
            shapiro2 = stats.shapiro(scores2)
            results['group2_normality'] = {
                'statistic': shapiro2.statistic,
                'p_value': shapiro2.pvalue,
                'is_normal': shapiro2.pvalue > 0.05
            }
        
        # Homogeneity of variance (Levene's test)
        if len(scores1) >= 2 and len(scores2) >= 2:
            levene = stats.levene(scores1, scores2)
            results['variance_homogeneity'] = {
                'statistic': levene.statistic,
                'p_value': levene.pvalue,
                'equal_variance': levene.pvalue > 0.05
            }
        
        return results

    def determine_best_variant(self) -> Dict:
        """Determine which variant performs best with statistical confidence"""
        results = []
        
        for variant_name in self.all_results:
            if variant_name == self.baseline_name:
                continue
                
            if "error" in self.all_results[variant_name]:
                continue
            
            analysis = self.calculate_distribution_shift(variant_name)
            consistency = self.calculate_consistency(variant_name)
            
            # Calculate statistical power
            baseline_scores = self.extract_scores(self.baseline_data)
            variant_scores = self.extract_scores(self.all_results[variant_name]["results"])
            
            power = self.calculate_statistical_power(
                effect_size=analysis.get('cohens_d', 0),
                n1=len(baseline_scores),
                n2=len(variant_scores)
            )
            
            # Apply multiple comparisons correction
            p_value = analysis.get('mann_whitney_pvalue', 1.0)
            
            results.append({
                'variant': variant_name,
                'mean_improvement': analysis.get('mean_delta', 0),
                'effect_size': analysis.get('cohens_d', 0),
                'p_value': p_value,
                'correlation': consistency.get('spearman_r', 0),
                'statistical_power': power,
                'has_score_1': analysis.get('variant_has_1s', False),
                'score': self._calculate_variant_score(analysis, consistency, power)
            })
        
        # Sort by composite score
        results.sort(key=lambda x: x['score'], reverse=True)
        
        # Apply FDR correction to all p-values
        p_values = [r['p_value'] for r in results]
        if p_values:
            _, significant = self.apply_fdr_correction(p_values)
            for r, sig in zip(results, significant):
                r['significant_after_correction'] = sig
        
        return {
            'best_variant': results[0] if results else None,
            'all_rankings': results,
            'recommendation': self._generate_recommendation(results)
        }

    def _calculate_variant_score(self, analysis: Dict, consistency: Dict, power: float) -> float:
        """Calculate composite score for ranking variants"""
        score = 0
        
        # Add safety check for baseline_data
        if self.baseline_data:
            baseline_scores = self.extract_scores(self.baseline_data)
            if not baseline_scores or 1 not in baseline_scores:
                if analysis.get('variant_has_1s', False):
                    score += 10
        else:
            # No baseline data, can't compare
            if analysis.get('variant_has_1s', False):
                score += 10
        
        # Reward statistical significance with good power
        if analysis.get('mann_whitney_pvalue', 1) < 0.05 and power > 0.8:
            score += 5
        
        # Reward high consistency
        score += consistency.get('spearman_r', 0) * 3
        
        # Consider effect size
        score += abs(analysis.get('cohens_d', 0)) * 2
        
        # Penalize very low power
        if power < 0.5:
            score *= 0.5
        
        return score



    def extract_scores(self, results: List[Dict], criteria_idx: Optional[int] = None) -> List[int]:
        """Extract scores from results - handles both string and int formats"""
        scores = []
        
        for result in results:
            if "case_study_1_llm" in result:
                cs1_raw = result["case_study_1_llm"]
                cs2_raw = result["case_study_2_llm"]
                
                # Handle both string and int formats
                cs1_scores = self._convert_scores(cs1_raw)
                cs2_scores = self._convert_scores(cs2_raw)
                
                if criteria_idx is not None:
                    # Get specific criteria
                    if len(cs1_scores) > criteria_idx:
                        scores.append(cs1_scores[criteria_idx])
                    if len(cs2_scores) > criteria_idx:
                        scores.append(cs2_scores[criteria_idx])
                else:
                    scores.extend(cs1_scores)
                    scores.extend(cs2_scores)
        
        return scores
    
    def _convert_scores(self, raw_scores: List[Union[str, int]]) -> List[int]:
        """Convert scores from string or int format to int list"""
        converted = []
        for s in raw_scores:
            if isinstance(s, str):
                if s not in ["00", "0"]:
                    try:
                        converted.append(int(s))
                    except ValueError:
                        logger.warning(f"Invalid score format: {s}")
            elif isinstance(s, (int, float)):
                if s != 0:
                    converted.append(int(s))
        return converted
    
    def validate_experiments(self) -> Dict[str, List[str]]:
        """Validate all experiments have sufficient data"""
        validation_results = {}
        
        for exp_name, exp_data in self.all_results.items():
            issues = []
            
            if "results" not in exp_data:
                issues.append("Missing results data")
                validation_results[exp_name] = issues
                continue
            
            results = exp_data["results"]
            
            if len(results) < self.MIN_SAMPLES:
                issues.append(f"Insufficient samples: {len(results)} < {self.MIN_SAMPLES}")
            
            scores = self.extract_scores(results)
            if len(scores) < 20:  # At least 20 scores needed for good stats
                issues.append(f"Insufficient scores: {len(scores)} < 20")
            
            # Check for variance
            if scores and np.std(scores) < 0.1:
                issues.append(f"Very low variance: std={np.std(scores):.3f}")
            
            # Check for data quality
            error_count = 0
            for r in results:
                if "00" in r.get("case_study_1_llm", []) or "00" in r.get("case_study_2_llm", []):
                    error_count += 1
            
            error_rate = error_count / len(results) if results else 0
            if error_rate > 0.2:
                issues.append(f"High error rate: {error_rate:.1%}")
            
            validation_results[exp_name] = issues
        
        return validation_results
    
    def calculate_distribution_shift_batch(self, variant_name: str, batch_size: int = 1000) -> Dict:
        """Calculate distribution shift with batching for large datasets"""

        # Add validation first
        if variant_name not in self.all_results:
            return {"error": f"Variant {variant_name} not found"}

        cache_key = f"dist_shift_{variant_name}"
        if cache_key in self._cache:
            return self._cache[cache_key]
        
        if not self.baseline_data:
            return {"error": "No baseline data"}
        
        # Process in batches to avoid memory issues
        baseline_scores = []
        variant_scores = []
        
        for i in range(0, len(self.baseline_data), batch_size):
            batch_baseline = self.baseline_data[i:i+batch_size]
            baseline_scores.extend(self.extract_scores(batch_baseline))
        
        variant_data = self.all_results[variant_name]["results"]
        for i in range(0, len(variant_data), batch_size):
            batch_variant = variant_data[i:i+batch_size]
            variant_scores.extend(self.extract_scores(batch_variant))
        
        result = self._calculate_distribution_metrics(baseline_scores, variant_scores, variant_name)
        self._add_to_cache(cache_key, result)
        return result
    
    def calculate_distribution_shift(self, variant_name: str) -> Dict:
        """Comprehensive distribution comparison"""
        # Use batch processing for large datasets
        if len(self.baseline_data) > 1000:
            return self.calculate_distribution_shift_batch(variant_name)
        
        if not self.baseline_data:
            return {"error": "No baseline data"}
            
        baseline_scores = self.extract_scores(self.baseline_data)
        variant_scores = self.extract_scores(self.all_results[variant_name]["results"])
        
        return self._calculate_distribution_metrics(baseline_scores, variant_scores, variant_name)
    
    def _safe_wilcoxon_test(self, baseline_paired: List, variant_paired: List):
        """Safely perform Wilcoxon test with proper validation"""
        if len(baseline_paired) < 5:  # Wilcoxon needs at least 5 pairs
            return None
        
        differences = [b - v for b, v in zip(baseline_paired, variant_paired)]
        non_zero_diff = [d for d in differences if d != 0]
        
        if len(non_zero_diff) < 1:
            return None
        
        try:
            return stats.wilcoxon(baseline_paired, variant_paired)
        except (ValueError, Exception) as e:
            logger.warning(f"Wilcoxon test failed: {e}")
            return None
    
    def _calculate_distribution_metrics(self, baseline_scores: List[int], variant_scores: List[int], variant_name: str) -> Dict:
        """Calculate distribution metrics between two score sets"""
        if len(baseline_scores) == 0 or len(variant_scores) == 0:
            return {"error": "Insufficient data"}
        
        baseline_dist = Counter(baseline_scores)
        variant_dist = Counter(variant_scores)
        
        # Calculate effect sizes
        cohens_d = self._calculate_cohens_d(baseline_scores, variant_scores)
        
        # KL Divergence
        kl_div = self._calculate_kl_divergence(baseline_dist, variant_dist)
        
        # Statistical tests with error handling
        try:
            ks_result = stats.ks_2samp(baseline_scores, variant_scores)
            mw_result = stats.mannwhitneyu(baseline_scores, variant_scores, alternative='two-sided')
        except Exception as e:
            logger.warning(f"Statistical test failed: {e}")
            ks_result = type('obj', (object,), {'statistic': np.nan, 'pvalue': np.nan})
            mw_result = type('obj', (object,), {'statistic': np.nan, 'pvalue': np.nan})
        
        # Wilcoxon test if paired data available - FIXED: now passes variant_name correctly
        wilcoxon_result = None
        paired_scores = self._get_paired_scores(variant_name)  # FIX: Pass variant name
        if len(paired_scores) > self.MIN_PAIRED_SAMPLES:
            baseline_paired = [p[0] for p in paired_scores]
            variant_paired = [p[1] for p in paired_scores]
            wilcoxon_result = self._safe_wilcoxon_test(baseline_paired, variant_paired)
        
        return {
            "mean_delta": np.mean(variant_scores) - np.mean(baseline_scores),
            "median_delta": np.median(variant_scores) - np.median(baseline_scores),
            "std_delta": np.std(variant_scores) - np.std(baseline_scores),
            "skew_change": stats.skew(variant_scores) - stats.skew(baseline_scores),
            "kurtosis_change": stats.kurtosis(variant_scores) - stats.kurtosis(baseline_scores),
            
            # Effect sizes
            "cohens_d": cohens_d,
            "effect_size_interpretation": self._interpret_cohens_d(cohens_d),
            
            # Score 1 analysis
            "baseline_has_1s": 1 in baseline_scores,
            "variant_has_1s": 1 in variant_scores,
            "score_1_count_delta": variant_dist.get(1, 0) - baseline_dist.get(1, 0),
            "score_1_pct_delta": (
                (variant_dist.get(1, 0)/len(variant_scores)*100 if variant_scores else 0) - 
                (baseline_dist.get(1, 0)/len(baseline_scores)*100 if baseline_scores else 0)
            ),
            
            # Score 5 analysis
            "score_5_count_delta": variant_dist.get(5, 0) - baseline_dist.get(5, 0),
            "score_5_pct_delta": (
                (variant_dist.get(5, 0)/len(variant_scores)*100 if variant_scores else 0) - 
                (baseline_dist.get(5, 0)/len(baseline_scores)*100 if baseline_scores else 0)
            ),
            
            # Distribution comparison
            "baseline_distribution": dict(baseline_dist),
            "variant_distribution": dict(variant_dist),
            "kl_divergence": kl_div,
            
            # Statistical tests
            "ks_statistic": ks_result.statistic,
            "ks_pvalue": ks_result.pvalue,
            "mann_whitney_u": mw_result.statistic,
            "mann_whitney_pvalue": mw_result.pvalue,
            "wilcoxon_statistic": wilcoxon_result.statistic if wilcoxon_result else None,
            "wilcoxon_pvalue": wilcoxon_result.pvalue if wilcoxon_result else None,
            
            # Sample sizes
            "baseline_n": len(baseline_scores),
            "variant_n": len(variant_scores)
        }
    
    def _calculate_cohens_d(self, group1: List[float], group2: List[float]) -> float:
        """Calculate Cohen's d effect size with better error handling"""
        if len(group1) == 0 or len(group2) == 0:
            return 0.0
            
        n1, n2 = len(group1), len(group2)
        
        # Need at least 2 samples for variance calculation
        if n1 < 2 or n2 < 2:
            return 0.0
        
        var1, var2 = np.var(group1, ddof=1), np.var(group2, ddof=1)
        
        # Pooled standard deviation
        pooled_var = ((n1 - 1) * var1 + (n2 - 1) * var2) / (n1 + n2 - 2)
        
        if pooled_var <= 0:
            return 0.0
        
        pooled_std = np.sqrt(pooled_var)
        
        if pooled_std < 1e-10:  # Better than == 0
            return 0.0  # FIX: Changed from 0.00 to 0.0
        
        return (np.mean(group2) - np.mean(group1)) / pooled_std
    
    def _interpret_cohens_d(self, d: float) -> str:
        """Interpret Cohen's d effect size"""
        abs_d = abs(d)
        if abs_d < 0.2:
            return "negligible"
        elif abs_d < 0.5:
            return "small"
        elif abs_d < 0.8:
            return "medium"
        else:
            return "large"
    
    def _calculate_kl_divergence(self, p_dist: Counter, q_dist: Counter) -> float:
        """Calculate KL divergence with smoothing and capping"""
        scores = [1, 2, 3, 4, 5]
        epsilon = 0.001  # Smoothing factor
        
        p = np.array([p_dist.get(s, 0) + epsilon for s in scores])
        q = np.array([q_dist.get(s, 0) + epsilon for s in scores])
        
        p = p / p.sum()
        q = q / q.sum()
        
        kl_div = stats.entropy(p, q)
        # Cap at reasonable maximum to avoid infinities
        return min(kl_div, 10.0)
    
    def _safe_correlation(self, x: List[float], y: List[float], method: str = 'pearson') -> Tuple[float, float]:
        """Calculate correlation with safety checks"""
        if len(x) < 2 or len(y) < 2:
            return np.nan, np.nan
        
        # Check for zero variance
        if np.std(x) < 1e-10 or np.std(y) < 1e-10:
            return np.nan, np.nan
        
        try:
            if method == 'pearson':
                return stats.pearsonr(x, y)
            elif method == 'spearman':
                return stats.spearmanr(x, y)
            elif method == 'kendall':
                return stats.kendalltau(x, y)
            else:
                return np.nan, np.nan
        except Exception as e:
            logger.warning(f"{method} correlation calculation failed: {e}")
            return np.nan, np.nan
    
    def calculate_consistency(self, variant_name: str) -> Dict:
        """Measure agreement between baseline and variant"""
        if not self.baseline_data:
            return {"error": "No baseline data"}
            
        paired_scores = self._get_paired_scores(variant_name)
        
        if not paired_scores:
            return {"error": "No paired scores found"}
        
        baseline_flat = [p[0] for p in paired_scores]
        variant_flat = [p[1] for p in paired_scores]
        
        exact_matches = sum(1 for b, v in paired_scores if b == v)
        direction_matches = sum(1 for b, v in paired_scores if (b > 3) == (v > 3))
        
        # Calculate correlation metrics with safety
        pearson_r, pearson_p = self._safe_correlation(baseline_flat, variant_flat, 'pearson')
        spearman_r, spearman_p = self._safe_correlation(baseline_flat, variant_flat, 'spearman')
        kendall_tau, kendall_p = self._safe_correlation(baseline_flat, variant_flat, 'kendall')
        
        # Calculate Cohen's Kappa if sklearn available
        cohen_kappa = None
        if HAS_SKLEARN:
            try:
                cohen_kappa = cohen_kappa_score(baseline_flat, variant_flat)
            except Exception as e:
                logger.warning(f"Cohen's kappa calculation failed: {e}")
        
        # Calculate ICC if pingouin available
        icc = self._calculate_icc(baseline_flat, variant_flat)
        
        return {
            "n_paired_scores": len(paired_scores),
            "exact_match_rate": exact_matches / len(paired_scores),
            "direction_match_rate": direction_matches / len(paired_scores),
            "mean_absolute_difference": np.mean([abs(b - v) for b, v in paired_scores]),
            "pearson_r": pearson_r,
            "pearson_p": pearson_p,
            "spearman_r": spearman_r,
            "spearman_p": spearman_p,
            "kendall_tau": kendall_tau,
            "kendall_p": kendall_p,
            "cohen_kappa": cohen_kappa,
            "icc": icc
        }
    
    def _calculate_icc(self, scores1: List[float], scores2: List[float]) -> Optional[float]:
        """Calculate Intraclass Correlation Coefficient"""
        if not HAS_PINGOUIN:
            return None
            
        try:
            df = pd.DataFrame({
                'targets': list(range(len(scores1))) * 2,
                'raters': [0] * len(scores1) + [1] * len(scores2),
                'scores': scores1 + scores2
            })
            icc_result = pg.intraclass_corr(data=df, targets='targets', 
                                           raters='raters', ratings='scores')
            return icc_result[icc_result['Type'] == 'ICC2']['ICC'].values[0]
        except Exception as e:
            logger.warning(f"ICC calculation failed: {e}")
            return None
    
    def _get_paired_scores(self, variant_name: str) -> List[Tuple[int, int]]:
        """Get paired scores with proper alignment"""
        paired_scores = []
        
        # Check variant exists
        if variant_name not in self.all_results:
            logger.warning(f"Variant {variant_name} not found in results")
            return paired_scores
        
        baseline_results = self.baseline_data
        variant_results = self.all_results[variant_name]["results"]
        
        # Create ID mapping
        baseline_by_id = {r["id"]: r for r in baseline_results}
        variant_by_id = {r["id"]: r for r in variant_results}
        
        # Find common IDs
        common_ids = sorted(set(baseline_by_id.keys()) & set(variant_by_id.keys()))
        
        for id_ in common_ids:
            b_result = baseline_by_id[id_]
            v_result = variant_by_id[id_]
            
            # Ensure we're pairing the same case studies and criteria
            for cs_key in ["case_study_1_llm", "case_study_2_llm"]:
                b_scores = b_result.get(cs_key, [])
                v_scores = v_result.get(cs_key, [])
                
                # Convert and validate scores
                b_valid = self._convert_scores(b_scores)
                v_valid = self._convert_scores(v_scores)
                
                # Only pair if both have same number of scores
                if len(b_valid) == len(v_valid) and len(b_valid) > 0:
                    for b_score, v_score in zip(b_valid, v_valid):
                        paired_scores.append((b_score, v_score))
                elif len(b_valid) != len(v_valid):
                    logger.warning(f"Score count mismatch for {id_} in {cs_key}: {len(b_valid)} vs {len(v_valid)}")
        
        return paired_scores
    
    def model_transformation(self, variant_name: str) -> Dict:
        """Model how variant transforms baseline scores"""
        paired = self._get_paired_scores(variant_name)
        
        if len(paired) < self.MIN_PAIRED_SAMPLES:
            return {"error": f"Insufficient paired data: {len(paired)} < {self.MIN_PAIRED_SAMPLES}"}
        
        baseline = np.array([p[0] for p in paired])
        variant = np.array([p[1] for p in paired])
        
        # Linear regression
        try:
            slope, intercept, r_value, p_value, std_err = stats.linregress(baseline, variant)
        except Exception as e:
            logger.warning(f"Linear regression failed: {e}")
            return {"error": f"Linear regression failed: {str(e)}"}
        
        # Polynomial regression (quadratic)
        try:
            poly_coeffs = np.polyfit(baseline, variant, 2)
            poly_r2 = self._calculate_r2(baseline, variant, poly_coeffs)
        except Exception as e:
            logger.warning(f"Polynomial regression failed: {e}")
            poly_coeffs = np.array([0, 0, 0])
            poly_r2 = 0
        
        # Create confusion matrix if sklearn available
        cm = None
        if HAS_SKLEARN:
            try:
                cm = confusion_matrix(baseline, variant, labels=[1, 2, 3, 4, 5])
            except Exception as e:
                logger.warning(f"Confusion matrix failed: {e}")
        
        # Interpret transformation
        interpretation = self._interpret_transformation(slope, intercept)
        
        return {
            "linear_model": {
                "slope": slope,
                "intercept": intercept,
                "r_squared": r_value**2,
                "p_value": p_value,
                "std_error": std_err,
                "equation": f"variant = {slope:.3f} * baseline + {intercept:.3f}"
            },
            "polynomial_model": {
                "coefficients": poly_coeffs.tolist(),
                "r_squared": poly_r2,
                "equation": f"variant = {poly_coeffs[0]:.3f}*baseline² + {poly_coeffs[1]:.3f}*baseline + {poly_coeffs[2]:.3f}"
            },
            "confusion_matrix": cm.tolist() if cm is not None else None,
            "transformation_type": interpretation,
            "residual_std": np.std(variant - (slope * baseline + intercept))
        }
    
    def _calculate_r2(self, x: np.ndarray, y: np.ndarray, coeffs: np.ndarray) -> float:
        """Calculate R² for polynomial fit"""
        y_pred = np.polyval(coeffs, x)
        ss_res = np.sum((y - y_pred) ** 2)
        ss_tot = np.sum((y - np.mean(y)) ** 2)
        return 1 - (ss_res / ss_tot) if ss_tot > 1e-10 else 0
    
    def _interpret_transformation(self, slope: float, intercept: float) -> str:
        """Interpret the transformation based on linear model"""
        interpretations = []
        
        if slope < 0.8:
            interpretations.append("compresses scores (reduces variance)")
        elif slope > 1.2:
            interpretations.append("expands scores (increases variance)")
        
        if intercept < -0.3:
            interpretations.append("systematic downward shift (more critical)")
        elif intercept > 0.3:
            interpretations.append("systematic upward shift (more lenient)")
        
        if not interpretations:
            return "minimal transformation"
        
        return " and ".join(interpretations)
    
    def identify_outliers(self, variant_name: str, n_outliers: int = 10) -> List[Dict]:
        """Find cases where experiments disagree most"""
        outliers = []
        
        if variant_name not in self.all_results:
            logger.warning(f"Variant {variant_name} not found")
            return outliers
        
        baseline_results = self.baseline_data
        variant_results = self.all_results[variant_name]["results"]
        
        # Create ID mapping
        baseline_by_id = {r["id"]: r for r in baseline_results}
        variant_by_id = {r["id"]: r for r in variant_results}
        
        common_ids = set(baseline_by_id.keys()) & set(variant_by_id.keys())
        
        for id_ in common_ids:
            b_result = baseline_by_id[id_]
            v_result = variant_by_id[id_]
            
            b_scores = self._extract_all_scores_from_result(b_result)
            v_scores = self._extract_all_scores_from_result(v_result)
            
            if b_scores and v_scores:
                b_mean = np.mean(b_scores)
                v_mean = np.mean(v_scores)
                delta = abs(b_mean - v_mean)
                
                outliers.append({
                    "id": id_,
                    "category": b_result.get("category_name", ""),
                    "baseline_mean": round(b_mean, 2),
                    "variant_mean": round(v_mean, 2),
                    "delta": round(delta, 2),
                    "baseline_scores": b_scores,
                    "variant_scores": v_scores,
                    "direction": "higher" if v_mean > b_mean else "lower"
                })
        
        # Sort by delta and return top N
        outliers.sort(key=lambda x: x["delta"], reverse=True)
        return outliers[:n_outliers]
    
    def _extract_all_scores_from_result(self, result: Dict) -> List[int]:
        """Extract all scores from a single result"""
        scores = []
        for key in ["case_study_1_llm", "case_study_2_llm"]:
            if key in result:
                scores.extend(self._convert_scores(result[key]))
        return scores
    
    def _analyze_criteria_effects(self, variant_name: str) -> Dict:
        """Analyze effects on individual criteria with proper mapping"""
        criteria_analysis = {}
        
        if variant_name not in self.all_results:
            logger.warning(f"Variant {variant_name} not found")
            return criteria_analysis
        
        for idx, criteria_name in self.CRITERIA_NAMES.items():
            baseline_scores = self.extract_scores(self.baseline_data, criteria_idx=idx)
            variant_scores = self.extract_scores(
                self.all_results[variant_name]["results"], 
                criteria_idx=idx
            )
            
            if len(baseline_scores) >= 10 and len(variant_scores) >= 10:
                # Add statistical test for each criterion
                try:
                    mw_result = stats.mannwhitneyu(baseline_scores, variant_scores, alternative='two-sided')
                    p_value = mw_result.pvalue
                except Exception as e:
                    logger.warning(f"Mann-Whitney U test failed for {criteria_name}: {e}")
                    p_value = np.nan
                
                criteria_analysis[criteria_name] = {
                    "baseline_mean": float(np.mean(baseline_scores)),
                    "variant_mean": float(np.mean(variant_scores)),
                    "delta": float(np.mean(variant_scores) - np.mean(baseline_scores)),
                    "effect_size": float(self._calculate_cohens_d(baseline_scores, variant_scores)),
                    "p_value": float(p_value),
                    "significant": bool(p_value < self.SIGNIFICANCE_LEVEL if not np.isnan(p_value) else False),
                    "n_baseline": len(baseline_scores),
                    "n_variant": len(variant_scores)
                }
            else:
                criteria_analysis[criteria_name] = {
                    "error": "Insufficient data",
                    "n_baseline": len(baseline_scores),
                    "n_variant": len(variant_scores)
                }
        
        return criteria_analysis

    def analyze_scoring_consistency(self, variant_name: str) -> Dict:
        """Analyze if variant is more/less consistent than baseline"""
        baseline_scores_by_id = {}
        variant_scores_by_id = {}
        
        # Group scores by ID to check within-ID variance
        for result in self.baseline_data:
            id_ = result["id"]
            scores = self.extract_scores([result])
            baseline_scores_by_id[id_] = scores
        
        for result in self.all_results[variant_name]["results"]:
            id_ = result["id"]
            scores = self.extract_scores([result])
            variant_scores_by_id[id_] = scores
        
        # Calculate within-ID variance
        baseline_variances = [np.var(scores) for scores in baseline_scores_by_id.values()]
        variant_variances = [np.var(scores) for scores in variant_scores_by_id.values()]
        
        return {
            "mean_within_id_variance_baseline": np.mean(baseline_variances),
            "mean_within_id_variance_variant": np.mean(variant_variances),
            "variance_increase": np.mean(variant_variances) - np.mean(baseline_variances),
            "more_consistent": np.mean(variant_variances) < np.mean(baseline_variances)
        }

    def analyze_category_effects(self, variant_name: str) -> Dict:
        """Which categories are most affected by the variant?"""
        category_impacts = {}
        
        # Discover categories from the data
        categories = set()
        for result in self.baseline_data:
            if "category_name" in result:
                categories.add(result["category_name"])
        
        for category in categories:
            baseline_cat = [r for r in self.baseline_data if r.get("category_name") == category]
            variant_cat = [r for r in self.all_results[variant_name]["results"] 
                        if r.get("category_name") == category]
            
            if baseline_cat and variant_cat:
                baseline_scores = self.extract_scores(baseline_cat)
                variant_scores = self.extract_scores(variant_cat)
                
                category_impacts[category] = {
                    "baseline_mean": np.mean(baseline_scores),
                    "variant_mean": np.mean(variant_scores),
                    "delta": np.mean(variant_scores) - np.mean(baseline_scores),
                    "effect_size": self._calculate_cohens_d(baseline_scores, variant_scores),
                    "sample_size": len(variant_cat)
                }
        
        # Rank by absolute effect size
        if category_impacts:
            ranked = sorted(category_impacts.items(), 
                        key=lambda x: abs(x[1]["effect_size"]), 
                        reverse=True)
            
            return {
                "category_impacts": category_impacts,
                "most_affected": ranked[:3] if len(ranked) >= 3 else ranked,
                "least_affected": ranked[-3:] if len(ranked) >= 3 else []
            }
        else:
            return {
                "category_impacts": {},
                "most_affected": [],
                "least_affected": []
            }

    def analyze_rationales(self, variant_name: str) -> Dict:
        """Analyze the rationales to understand scoring differences"""
        from collections import Counter
        import re
        import numpy as np  # Added missing import
        
        # Collect paired rationales and scores
        baseline_rationales = []
        baseline_scores = []
        variant_rationales = []
        variant_scores = []
        
        def collect_rationales_scores(data_list, rationales_list, scores_list):
            """Helper function to collect rationales and scores"""
            for result in data_list:
                for i, cs_rat in enumerate(["case_study_1_rationales", "case_study_2_rationales"]):
                    cs_scores = ["case_study_1_llm", "case_study_2_llm"][i]
                    
                    if cs_rat in result and cs_scores in result:
                        rationales = result[cs_rat]
                        scores = self._convert_scores(result[cs_scores])
                        
                        # Handle both dict and list types for rationales
                        if isinstance(rationales, dict):
                            # Ensure we're iterating in a consistent order
                            for criteria_name in sorted(rationales.keys()):
                                rationale = rationales[criteria_name]
                                # Find the corresponding score index
                                criteria_keys = sorted(rationales.keys())
                                criteria_idx = criteria_keys.index(criteria_name)
                                
                                if criteria_idx < len(scores):
                                    rationales_list.append(rationale if rationale else "")
                                    scores_list.append(scores[criteria_idx])
                        elif isinstance(rationales, list):
                            # If rationales is already a list
                            for idx, rationale in enumerate(rationales):
                                if idx < len(scores):
                                    rationales_list.append(rationale if rationale else "")
                                    scores_list.append(scores[idx])
        
        # Collect baseline data
        collect_rationales_scores(self.baseline_data, baseline_rationales, baseline_scores)
        
        # Collect variant data
        if variant_name in self.all_results and "results" in self.all_results[variant_name]:
            collect_rationales_scores(
                self.all_results[variant_name]["results"], 
                variant_rationales, 
                variant_scores
            )
        
        # Common phrases/words in low scores vs high scores
        def extract_keywords(rationales, scores, min_word_length=3):
            """Extract keywords from rationales based on score ranges"""
            low_score_words = []
            high_score_words = []
            
            # Common stop words to filter out
            stop_words = {'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 
                        'of', 'with', 'by', 'from', 'is', 'was', 'are', 'were', 'been', 'be',
                        'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could',
                        'should', 'may', 'might', 'can', 'this', 'that', 'these', 'those'}
            
            for rat, score in zip(rationales, scores):
                if not rat:  # Skip empty rationales
                    continue
                    
                # Extract words, filtering short ones and stop words
                words = [w for w in re.findall(r'\b\w+\b', rat.lower()) 
                        if len(w) >= min_word_length and w not in stop_words]
                
                if score <= 2:
                    low_score_words.extend(words)
                elif score >= 4:
                    high_score_words.extend(words)
            
            return Counter(low_score_words), Counter(high_score_words)
        
        # Extract keywords
        baseline_low_words, baseline_high_words = extract_keywords(baseline_rationales, baseline_scores)
        variant_low_words, variant_high_words = extract_keywords(variant_rationales, variant_scores)
        
        # Get top keywords
        top_n = 15  # Increased from 10 for better insight
        
        # Length analysis (excluding empty rationales)
        baseline_lengths = [len(r) for r in baseline_rationales if r]
        variant_lengths = [len(r) for r in variant_rationales if r]
        
        # Calculate statistics
        results = {
            "avg_rationale_length_baseline": np.mean(baseline_lengths) if baseline_lengths else 0,
            "avg_rationale_length_variant": np.mean(variant_lengths) if variant_lengths else 0,
            "length_change": (np.mean(variant_lengths) - np.mean(baseline_lengths)) if baseline_lengths and variant_lengths else 0,
            "total_rationales_baseline": len(baseline_rationales),
            "total_rationales_variant": len(variant_rationales),
            "empty_rationales_baseline": sum(1 for r in baseline_rationales if not r),
            "empty_rationales_variant": sum(1 for r in variant_rationales if not r),
            "baseline_low_score_keywords": dict(baseline_low_words.most_common(top_n)),
            "baseline_high_score_keywords": dict(baseline_high_words.most_common(top_n)),
            "variant_low_score_keywords": dict(variant_low_words.most_common(top_n)),
            "variant_high_score_keywords": dict(variant_high_words.most_common(top_n)),
        }
        
        # Add score distribution analysis
        if baseline_scores and variant_scores:
            results.update({
                "avg_score_baseline": np.mean(baseline_scores),
                "avg_score_variant": np.mean(variant_scores),
                "score_change": np.mean(variant_scores) - np.mean(baseline_scores),
                "std_score_baseline": np.std(baseline_scores),
                "std_score_variant": np.std(variant_scores),
            })
        
        return results

    def analyze_score_flips(self, variant_name: str) -> Dict:
        """Find where scores flipped from pass to fail or vice versa"""
        paired = self._get_paired_scores(variant_name)
        
        flips = {
            "fail_to_pass": [],  # baseline <=2, variant >=3
            "pass_to_fail": [],  # baseline >=3, variant <=2
            "major_drops": [],   # dropped by 2+ points
            "major_gains": []    # gained by 2+ points
        }
        
        for baseline_score, variant_score in paired:
            delta = variant_score - baseline_score
            
            if baseline_score <= 2 and variant_score >= 3:
                flips["fail_to_pass"].append((baseline_score, variant_score))
            elif baseline_score >= 3 and variant_score <= 2:
                flips["pass_to_fail"].append((baseline_score, variant_score))
            
            if delta <= -2:
                flips["major_drops"].append((baseline_score, variant_score))
            elif delta >= 2:
                flips["major_gains"].append((baseline_score, variant_score))
        
        return {
            "fail_to_pass_count": len(flips["fail_to_pass"]),
            "pass_to_fail_count": len(flips["pass_to_fail"]),
            "major_drops_count": len(flips["major_drops"]),
            "major_gains_count": len(flips["major_gains"]),
            "net_flip_impact": len(flips["fail_to_pass"]) - len(flips["pass_to_fail"]),
            "examples": {k: v[:5] for k, v in flips.items()}  # First 5 examples
        }
        
    def analyze_referee_alignment(self, variant_name: str) -> Dict:
        """How well do scores align with referee assessments?"""
        def get_referee_score(result):
            """Convert referee responses to a score"""
            referee_scores = []
            
            # Check both case studies
            for cs_num in [1, 2]:
                for i in range(1, 6):
                    ref_key = f"referee_{i}_cs{cs_num}"
                    val = result.get(ref_key)
                    if val is not None:
                        # Add type checking
                        if isinstance(val, str):
                            if val == "Yes":
                                referee_scores.append(5)
                            elif val == "No":
                                referee_scores.append(1)
                            elif val == "Partially":
                                referee_scores.append(3)
                            elif val != "N/A":
                                # Try to parse if it's a different format
                                try:
                                    val_lower = val.lower()
                                    if "yes" in val_lower:
                                        referee_scores.append(5)
                                    elif "no" in val_lower:
                                        referee_scores.append(1)
                                    elif "partial" in val_lower:
                                        referee_scores.append(3)
                                except AttributeError:
                                    pass  # val might not have .lower() method
            
            return np.mean(referee_scores) if referee_scores else None

        def calculate_alignment(llm_scores, referee_score):
            """Calculate alignment between LLM and referee scores"""
            if not llm_scores or referee_score is None:
                return None
            
            llm_mean = np.mean(llm_scores)
            # Alignment is inverse of distance (lower is better)
            return abs(llm_mean - referee_score)
        
        baseline_alignments = []
        variant_alignments = []
        baseline_correlations = []
        variant_correlations = []
        
        # Analyze baseline alignment
        for result in self.baseline_data:
            ref_score = get_referee_score(result)
            if ref_score is not None:
                llm_scores = self.extract_scores([result])
                if llm_scores:
                    alignment = calculate_alignment(llm_scores, ref_score)
                    if alignment is not None:
                        baseline_alignments.append(alignment)
                        baseline_correlations.append((ref_score, np.mean(llm_scores)))
        
        # Analyze variant alignment
        if variant_name in self.all_results:
            for result in self.all_results[variant_name]["results"]:
                ref_score = get_referee_score(result)
                if ref_score is not None:
                    llm_scores = self.extract_scores([result])
                    if llm_scores:
                        alignment = calculate_alignment(llm_scores, ref_score)
                        if alignment is not None:
                            variant_alignments.append(alignment)
                            variant_correlations.append((ref_score, np.mean(llm_scores)))
        
        # Calculate correlation coefficients if we have enough data
        baseline_correlation = np.nan
        variant_correlation = np.nan
        
        if len(baseline_correlations) >= 10:
            refs, llms = zip(*baseline_correlations)
            baseline_correlation, _ = self._safe_correlation(refs, llms, 'pearson')
        
        if len(variant_correlations) >= 10:
            refs, llms = zip(*variant_correlations)
            variant_correlation, _ = self._safe_correlation(refs, llms, 'pearson')
        
        # Calculate summary statistics
        baseline_mean_alignment = np.mean(baseline_alignments) if baseline_alignments else np.nan
        variant_mean_alignment = np.mean(variant_alignments) if variant_alignments else np.nan
        
        # Perfect alignment would be 0 (no difference), so lower is better
        alignment_improved = variant_mean_alignment < baseline_mean_alignment if not np.isnan(variant_mean_alignment) and not np.isnan(baseline_mean_alignment) else None
        
        # Higher correlation is better
        correlation_improved = variant_correlation > baseline_correlation if not np.isnan(variant_correlation) and not np.isnan(baseline_correlation) else None
        
        return {
            "baseline_referee_alignment": baseline_mean_alignment,
            "variant_referee_alignment": variant_mean_alignment,
            "alignment_difference": variant_mean_alignment - baseline_mean_alignment if not np.isnan(variant_mean_alignment) and not np.isnan(baseline_mean_alignment) else np.nan,
            "alignment_improved": alignment_improved,
            
            "baseline_correlation": baseline_correlation,
            "variant_correlation": variant_correlation,
            "correlation_improved": correlation_improved,
            
            "baseline_samples_with_referees": len(baseline_alignments),
            "variant_samples_with_referees": len(variant_alignments),
            
            "interpretation": self._interpret_referee_alignment(
                baseline_mean_alignment, 
                variant_mean_alignment,
                baseline_correlation,
                variant_correlation
            )
        }

    def _interpret_referee_alignment(self, baseline_align: float, variant_align: float, 
                                    baseline_corr: float, variant_corr: float) -> str:
        """Interpret referee alignment results"""
        interpretations = []
        
        # Alignment interpretation (lower is better)
        if not np.isnan(baseline_align) and not np.isnan(variant_align):
            if variant_align < baseline_align - 0.2:
                interpretations.append("significantly better alignment with referees")
            elif variant_align < baseline_align:
                interpretations.append("slightly better alignment with referees")
            elif variant_align > baseline_align + 0.2:
                interpretations.append("significantly worse alignment with referees")
            elif variant_align > baseline_align:
                interpretations.append("slightly worse alignment with referees")
            else:
                interpretations.append("similar alignment with referees")
        
        # Correlation interpretation (higher is better)
        if not np.isnan(baseline_corr) and not np.isnan(variant_corr):
            if variant_corr > baseline_corr + 0.1:
                interpretations.append("stronger correlation with referee scores")
            elif variant_corr < baseline_corr - 0.1:
                interpretations.append("weaker correlation with referee scores")
        
        if not interpretations:
            return "insufficient referee data for comparison"
        
        return " and ".join(interpretations)

    def analyze_context_effects(self, variant_name: str) -> Dict:
        """Compare Capability vs Capacity context performance"""
        contexts = {}
        
        for context in ["Capability", "Capacity"]:
            baseline_context = []
            variant_context = []
            
            for result in self.baseline_data:
                if result.get("context_1") == context:
                    scores = result.get("case_study_1_llm", [])
                    baseline_context.extend(self._convert_scores(scores))
                if result.get("context_2") == context:
                    scores = result.get("case_study_2_llm", [])
                    baseline_context.extend(self._convert_scores(scores))
            
            # Similar for variant...
            if variant_name in self.all_results:
                for result in self.all_results[variant_name]["results"]:
                    if result.get("context_1") == context:
                        scores = result.get("case_study_1_llm", [])
                        variant_context.extend(self._convert_scores(scores))
                    if result.get("context_2") == context:
                        scores = result.get("case_study_2_llm", [])
                        variant_context.extend(self._convert_scores(scores))

            contexts[context] = {
                "baseline_mean": np.mean(baseline_context),
                "variant_mean": np.mean(variant_context),
                "delta": np.mean(variant_context) - np.mean(baseline_context),
                "effect_size": self._calculate_cohens_d(baseline_context, variant_context)
            }
        
        return contexts

    def generate_recommendations(self) -> List[Dict]:
        """Generate actionable recommendations"""
        recommendations = []
        
        if not self.baseline_data:
            return recommendations
        
        baseline_scores = self.extract_scores(self.baseline_data)
        if not baseline_scores:
            return recommendations
            
        baseline_mean = np.mean(baseline_scores)
        baseline_has_1s = 1 in baseline_scores
        
        for variant_name, variant_data in self.all_results.items():
            if variant_name == self.baseline_name:
                continue
            
            if "error" in variant_data:
                continue
            
            analysis = {
                "distribution": self.calculate_distribution_shift(variant_name),
                "consistency": self.calculate_consistency(variant_name),
                "transformation": self.model_transformation(variant_name)
            }
            
            # Skip if analysis failed
            if "error" in analysis["distribution"]:
                continue
            
            # Fix missing score 1 problem
            if not baseline_has_1s and analysis["distribution"]["variant_has_1s"]:
                recommendations.append({
                    "experiment": variant_name,
                    "priority": "HIGH",
                    "category": "distribution_fix",
                    "reason": "Fixes missing score 1 problem",
                    "details": f"Introduces {analysis['distribution']['score_1_count_delta']} score 1s",
                    "metrics": {
                        "mean_shift": analysis["distribution"]["mean_delta"],
                        "correlation": analysis["consistency"].get("spearman_r", 0),
                        "effect_size": analysis["distribution"]["cohens_d"]
                    },
                    "statistical_significance": analysis["distribution"]["mann_whitney_pvalue"] < self.SIGNIFICANCE_LEVEL
                })
            
            # Fix overscoring
            if baseline_mean > 3.5 and analysis["distribution"]["mean_delta"] < -0.3:
                recommendations.append({
                    "experiment": variant_name,
                    "priority": "MEDIUM",
                    "category": "calibration",
                    "reason": "Reduces overscoring tendency",
                    "details": f"Lowers mean from {baseline_mean:.2f} to {baseline_mean + analysis['distribution']['mean_delta']:.2f}",
                    "metrics": {
                        "mean_shift": analysis["distribution"]["mean_delta"],
                        "correlation": analysis["consistency"].get("spearman_r", 0),
                        "effect_size": analysis["distribution"]["cohens_d"]
                    },
                    "statistical_significance": analysis["distribution"]["mann_whitney_pvalue"] < self.SIGNIFICANCE_LEVEL
                })
            
            # High consistency alternative
            if (analysis["consistency"].get("spearman_r", 0) > 0.8 and 
                abs(analysis["distribution"]["mean_delta"]) < 0.2):
                recommendations.append({
                    "experiment": variant_name,
                    "priority": "LOW",
                    "category": "alternative",
                    "reason": "Highly consistent alternative with minimal impact",
                    "details": f"Correlation: {analysis['consistency']['spearman_r']:.3f}, Mean shift: {analysis['distribution']['mean_delta']:.2f}",
                    "metrics": {
                        "mean_shift": analysis["distribution"]["mean_delta"],
                        "correlation": analysis["consistency"].get("spearman_r", 0),
                        "effect_size": analysis["distribution"]["cohens_d"]
                    },
                    "statistical_significance": False
                })
        
        # Sort by priority and correlation
        priority_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
        recommendations.sort(
            key=lambda x: (
                priority_order.get(x["priority"], 3),
                -x["metrics"]["correlation"]
            )
        )
        
        return recommendations
    
    def _safe_get_numeric(self, value: Any, default: float = 0.0) -> float:
        """Safely convert to numeric value"""
        if value is None:
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default
    
    def _get_basic_stats(self, results: List[Dict]) -> Dict:
        """Get basic statistics for a result set"""
        scores = self.extract_scores(results)
        
        if len(scores) == 0:
            return {
                "mean_score": 0,
                "median_score": 0,
                "std_dev": 0,
                "has_score_1": False,
                "min_score": 0,
                "max_score": 0,
                "total_scores": 0,
                "score_distribution": {},
                "score_1_percentage": 0,
                "skewness": 0,
                "kurtosis": 0
            }
        
        return {
            "mean_score": float(np.mean(scores)),
            "median_score": float(np.median(scores)),
            "std_dev": float(np.std(scores)),
            "min_score": int(min(scores)),
            "max_score": int(max(scores)),
            "total_scores": int(len(scores)),
            "score_distribution": {int(k): int(v) for k, v in dict(Counter(scores)).items()},
            "has_score_1": bool(1 in scores),
            "score_1_percentage": float(scores.count(1) / len(scores) * 100),
            "has_score_5": bool(5 in scores),
            "score_5_percentage": float(scores.count(5) / len(scores) * 100),
            "skewness": float(stats.skew(scores)),
            "kurtosis": float(stats.kurtosis(scores))
        }
    
    def _get_config_diff(self, variant_name: str) -> Dict:
        """Get configuration differences from baseline"""
        if self.baseline_name not in self.all_results or variant_name not in self.all_results:
            return {}
            
        baseline_config = self.all_results[self.baseline_name].get("config", {})
        variant_config = self.all_results[variant_name].get("config", {})
        
        # Handle both object and dict configs
        if hasattr(baseline_config, "config"):
            baseline_config = baseline_config.config
        if hasattr(variant_config, "config"):
            variant_config = variant_config.config
        
        diff = {}
        for key in set(baseline_config.keys()) | set(variant_config.keys()):
            baseline_val = baseline_config.get(key)
            variant_val = variant_config.get(key)
            if baseline_val != variant_val:
                diff[key] = {
                    "baseline": baseline_val,
                    "variant": variant_val
                }
        
        return diff
    
    def export_analysis_config(self, output_dir: str):
        """Export the analysis configuration for reproducibility"""
        config = {
            "analysis_version": "1.1.0",
            "timestamp": datetime.now().isoformat(),
            "baseline_experiment": self.baseline_name,
            "experiments_analyzed": list(self.all_results.keys()),
            "statistical_tests": [
                "mann_whitney_u",
                "kolmogorov_smirnov", 
                "wilcoxon_signed_rank",
                "cohens_d",
                "kl_divergence",
                "pearson_correlation",
                "spearman_correlation",
                "kendall_tau"
            ],
            "criteria_names": self.CRITERIA_NAMES,
            "thresholds": {
                "significance_level": self.SIGNIFICANCE_LEVEL,
                "min_samples": self.MIN_SAMPLES,
                "min_paired_samples": self.MIN_PAIRED_SAMPLES,
                "max_cache_size": self.MAX_CACHE_SIZE
            }
        }
        
        config_file = os.path.join(output_dir, "analysis_config.json")
        with open(config_file, 'w') as f:
            json.dump(config, f, indent=2, cls=NumpyJSONEncoder)

    def generate_failure_analysis_report(self, output_dir: str):
        """Generate detailed failure analysis"""
        
        report = {
            "timestamp": datetime.now().isoformat(),
            "baseline": self.baseline_name,
            "experiments_analyzed": list(self.all_results.keys())
        }
        
        for variant_name in self.all_results:
            if variant_name == self.baseline_name:
                continue
            
            if "error" in self.all_results[variant_name]:
                continue
                
            report[variant_name] = {
                "consistency": self.analyze_scoring_consistency(variant_name),
                "category_effects": self.analyze_category_effects(variant_name),
                "score_flips": self.analyze_score_flips(variant_name),
                "context_effects": self.analyze_context_effects(variant_name),
                "rationales": self.analyze_rationales(variant_name),
                "referee_alignment": self.analyze_referee_alignment(variant_name)
            }
        
        failure_file = os.path.join(output_dir, "failure_analysis.json")
        with open(failure_file, 'w') as f:
            json.dump(report, f, indent=2, cls=NumpyJSONEncoder)
        
        logger.info(f"Failure analysis written to {failure_file}")

    def generate_all_reports(self, output_root: str):
        """Generate comprehensive analysis reports with progress tracking"""
        
        # Create consolidated output directory
        consolidated_dir = os.path.join(output_root, "consolidated")
        os.makedirs(consolidated_dir, exist_ok=True)
        
        logger.info("="*60)
        logger.info("Starting statistical analysis...")
        logger.info(f"Analyzing {len(self.all_results)} experiments")
        logger.info(f"Baseline: {self.baseline_name}")
        logger.info("="*60)
        
        # Validate experiments first
        logger.info("Validating experiments...")
        validation = self.validate_experiments()
        validation_passed = True
        for exp_name, issues in validation.items():
            if issues:
                logger.warning(f"Validation issues for {exp_name}: {', '.join(issues)}")
                validation_passed = False
        
        if validation_passed:
            logger.info("✓ All experiments passed validation")
        
        # Save validation results
        with open(os.path.join(consolidated_dir, "validation_results.json"), 'w') as f:
            json.dump(validation, f, indent=2, cls=NumpyJSONEncoder)
        
        # Generate reports
        steps = [
            ("individual statistics", lambda: self._generate_individual_stats(output_root)),
            ("comparative analysis", lambda: self._generate_comparative_analysis()),
            ("distribution CSV", lambda: self._generate_distribution_csv(consolidated_dir)),
            ("recommendations", lambda: self.generate_recommendations()),
            ("analysis configuration", lambda: self.export_analysis_config(consolidated_dir)),
            ("failure analysis", lambda: self.generate_failure_analysis_report(consolidated_dir))
        ]
        
        comparative = None
        recommendations = None
        
        for step_name, step_func in steps:
            logger.info(f"Generating {step_name}...")
            try:
                result = step_func()
                
                if step_name == "comparative analysis":
                    comparative = result
                    with open(os.path.join(consolidated_dir, "comparative_analysis.json"), 'w') as f:
                        json.dump(comparative, f, indent=2, cls=NumpyJSONEncoder, default=str)
                elif step_name == "recommendations":
                    recommendations = result
                    with open(os.path.join(consolidated_dir, "recommendations.json"), 'w') as f:
                        json.dump(recommendations, f, indent=2, cls=NumpyJSONEncoder)
            except Exception as e:
                logger.error(f"Failed to generate {step_name}: {e}")
        
        # Generate executive summary
        logger.info("Generating executive summary...")
        try:
            self._generate_executive_summary(consolidated_dir, comparative, recommendations)
        except Exception as e:
            logger.error(f"Failed to generate executive summary: {e}")
        
        # Generate Excel report
        logger.info("Generating Excel report...")
        try:
            self._generate_excel_report(consolidated_dir, comparative, recommendations)
        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")
        
        # Generate visualizations
        try:
            logger.info("Generating visualizations...")
            self._generate_visualizations(consolidated_dir)
        except ImportError:
            logger.warning("Visualization libraries not available")
        except Exception as e:
            logger.warning(f"Failed to generate visualizations: {e}")
        
        logger.info("="*60)
        logger.info(f"✓ Reports generated successfully in {consolidated_dir}")
        logger.info("="*60)
    

    def _generate_individual_stats(self, output_root: str):
        """Generate statistics for each experiment"""
        for exp_name, exp_data in self.all_results.items():
            exp_dir = os.path.join(output_root, exp_name)
            
            if "error" in exp_data:
                continue
                
            scores = self.extract_scores(exp_data["results"])
            
            # Fix unsafe config access
            config = exp_data.get("config")
            stats = {
                "experiment_name": config.name if config and hasattr(config, "name") else exp_name,
                "configuration": config.to_dict() if config and hasattr(config, "to_dict") else {},
                "basic_stats": self._get_basic_stats(exp_data["results"]),
                "criteria_breakdown": self._analyze_criteria_effects(exp_name)
                    if exp_name != self.baseline_name else {},
                "timestamp": datetime.now().isoformat()
            }
            
            stats_file = os.path.join(exp_dir, "stats.json")
            with open(stats_file, 'w') as f:
                # Fix missing space after comma
                json.dump(stats, f, indent=2, cls=NumpyJSONEncoder)

    def _generate_comparative_analysis(self) -> Dict:
        """Generate full comparative analysis"""
        analysis = {
            "baseline": self.baseline_name,
            "baseline_stats": self._get_basic_stats(self.baseline_data) if self.baseline_data else {},
            "variants": {},
            "timestamp": datetime.now().isoformat()
        }
        
        for variant_name in self.all_results:
            if variant_name == self.baseline_name:
                continue
            
            if "error" in self.all_results[variant_name]:
                analysis["variants"][variant_name] = {
                    "error": self.all_results[variant_name]["error"]
                }
                continue
            
            # Get scores for new analyses
            baseline_scores = self.extract_scores(self.baseline_data)
            variant_scores = self.extract_scores(self.all_results[variant_name]["results"])
            
            # NEW: Add advanced statistics
            advanced_stats = {}
            
            # Cliff's Delta (better for ordinal data than Cohen's d)
            cliffs_delta, cliffs_interpretation = self.calculate_cliffs_delta(baseline_scores, variant_scores)
            advanced_stats["cliffs_delta"] = {
                "value": cliffs_delta,
                "interpretation": cliffs_interpretation
            }
            
            # Statistical power analysis
            effect_size = self._calculate_cohens_d(baseline_scores, variant_scores)
            power = self.calculate_statistical_power(
                effect_size=effect_size,
                n1=len(baseline_scores),
                n2=len(variant_scores)
            )
            required_n = self.required_sample_size(effect_size) if abs(effect_size) > 0.1 else None
            advanced_stats["power_analysis"] = {
                "statistical_power": power,
                "required_sample_size": required_n,
                "current_sample_size": len(baseline_scores),
                "is_adequately_powered": power > 0.8
            }
            
            # Assumption testing
            advanced_stats["assumption_tests"] = self.test_assumptions(baseline_scores, variant_scores)
            
            # Bayesian analysis
            advanced_stats["bayesian"] = self.bayesian_comparison(baseline_scores, variant_scores)
            
            # Cohen's d with confidence interval
            d, ci_lower, ci_upper = self.calculate_cohens_d_ci(baseline_scores, variant_scores)
            advanced_stats["cohens_d_ci"] = {
                "estimate": d,
                "ci_95_lower": ci_lower,
                "ci_95_upper": ci_upper
            }
            
            # Practical significance
            advanced_stats["practical_significance"] = self.analyze_practical_significance(variant_name)
            
            # Krippendorff's alpha for reliability
            paired = self._get_paired_scores(variant_name)
            if len(paired) > 10:
                baseline_paired = [p[0] for p in paired]
                variant_paired = [p[1] for p in paired]
                advanced_stats["krippendorff_alpha"] = self.calculate_krippendorff_alpha(
                    baseline_paired, variant_paired
                )
            
            # Add everything to the analysis
            analysis["variants"][variant_name] = {
                "config_diff": self._get_config_diff(variant_name),
                "distribution_shift": self.calculate_distribution_shift(variant_name),
                "consistency": self.calculate_consistency(variant_name),
                "transformation": self.model_transformation(variant_name),
                "top_outliers": self.identify_outliers(variant_name, n_outliers=5),
                "criteria_analysis": self._analyze_criteria_effects(variant_name),
                "scoring_consistency": self.analyze_scoring_consistency(variant_name),
                "category_effects": self.analyze_category_effects(variant_name),
                "score_flips": self.analyze_score_flips(variant_name),
                "context_effects": self.analyze_context_effects(variant_name),
                "rationale_analysis": self.analyze_rationales(variant_name),
                "referee_alignment": self.analyze_referee_alignment(variant_name),
                "advanced_statistics": advanced_stats  # NEW
            }
        
        # NEW: Add overall best variant determination
        analysis["best_variant_analysis"] = self.determine_best_variant()
        
        return analysis


    def _generate_distribution_csv(self, output_dir: str):
        """Generate CSV with score distributions"""
        rows = []
        
        for exp_name, exp_data in self.all_results.items():
            if "error" in exp_data:
                continue
                
            scores = self.extract_scores(exp_data["results"])
            if not scores:
                continue
                
            dist = Counter(scores)
            
            row = {
                "experiment": exp_name,
                "is_baseline": exp_name == self.baseline_name,
                "mean": np.mean(scores),
                "median": np.median(scores),
                "std": np.std(scores),
                "n": len(scores),
                "score_1": dist.get(1, 0),
                "score_2": dist.get(2, 0),
                "score_3": dist.get(3, 0),
                "score_4": dist.get(4, 0),
                "score_5": dist.get(5, 0),
                "pct_1": dist.get(1, 0) / len(scores) * 100,
                "pct_2": dist.get(2, 0) / len(scores) * 100,
                "pct_3": dist.get(3, 0) / len(scores) * 100,
                "pct_4": dist.get(4, 0) / len(scores) * 100,
                "pct_5": dist.get(5, 0) / len(scores) * 100
            }
            rows.append(row)
        
        if rows:
            df = pd.DataFrame(rows)
            df.to_csv(os.path.join(output_dir, "score_distributions.csv"), index=False)

    def _create_advanced_statistics_dataframe(self, comparative: Dict) -> Optional[pd.DataFrame]:
        """Create advanced statistics dataframe for Excel"""
        if not comparative:
            return None
        
        rows = []
        
        for variant_name, variant_data in comparative.get("variants", {}).items():
            if "error" in variant_data or "advanced_statistics" not in variant_data:
                continue
            adv = variant_data["advanced_statistics"]
            
            row = {
                "Experiment": variant_name,
                "Cliff's Delta": adv.get("cliffs_delta", {}).get("value", np.nan),
                "Cliff's Interpretation": adv.get("cliffs_delta", {}).get("interpretation", ""),
                "Cohen's d": adv.get("cohens_d_ci", {}).get("estimate", np.nan),
                "Cohen's d CI Lower": adv.get("cohens_d_ci", {}).get("ci_95_lower", np.nan),
                "Cohen's d CI Upper": adv.get("cohens_d_ci", {}).get("ci_95_upper", np.nan),
                "Statistical Power": adv.get("power_analysis", {}).get("statistical_power", np.nan),
                "Required N": adv.get("power_analysis", {}).get("required_sample_size", np.nan),
                "Adequately Powered": adv.get("power_analysis", {}).get("is_adequately_powered", False),
                "Bayesian P(Better)": adv.get("bayesian", {}).get("prob_variant_better", np.nan),
                "Bayesian Interpretation": adv.get("bayesian", {}).get("interpretation", ""),
                "Practically Significant": adv.get("practical_significance", {}).get("is_practically_significant", False),
                "NNT": adv.get("practical_significance", {}).get("nnt_for_success", np.nan),
                "Krippendorff Alpha": adv.get("krippendorff_alpha", np.nan),
                "Data Normal": adv.get("assumption_tests", {}).get("group1_normality", {}).get("is_normal", False),
                "Equal Variance": adv.get("assumption_tests", {}).get("variance_homogeneity", {}).get("equal_variance", False)
            }
            rows.append(row)
        
        return pd.DataFrame(rows) if rows else None

    def _generate_excel_report(self, output_dir: str, comparative: Dict, recommendations: List):
        """Generate comprehensive Excel report with multiple sheets"""

        # Handle None inputs
        if comparative is None:
            logger.error("Cannot generate Excel report: comparative analysis failed")
            return
        if recommendations is None:
            recommendations = []

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.warning("openpyxl not installed, trying xlsxwriter")
            try:
                excel_file = os.path.join(output_dir, "analysis_report.xlsx")
                with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
                    self._write_excel_sheets_basic(writer, comparative, recommendations)
                logger.info(f"Excel report saved to {excel_file}")
                return
            except ImportError:
                logger.error("Neither openpyxl nor xlsxwriter available for Excel export")
                return
        
        excel_file = os.path.join(output_dir, "analysis_report.xlsx")
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 1. Summary Sheet
            summary_data = self._create_summary_dataframe(comparative)
            if not summary_data.empty:
                summary_data.to_excel(writer, sheet_name='Summary', index=False)
            
            # 2. Score Distributions
            dist_data = self._create_distribution_dataframe()
            if not dist_data.empty:
                dist_data.to_excel(writer, sheet_name='Score Distributions', index=False)
            
            # 3. Statistical Tests
            stats_data = self._create_statistical_tests_dataframe(comparative)
            if not stats_data.empty:
                stats_data.to_excel(writer, sheet_name='Statistical Tests', index=False)
            
            # 4. Recommendations
            if recommendations:
                rec_data = pd.DataFrame(recommendations)
                rec_data.to_excel(writer, sheet_name='Recommendations', index=False)
            
            # 5. Criteria Analysis
            criteria_data = self._create_criteria_analysis_dataframe(comparative)
            if criteria_data is not None and not criteria_data.empty:
                criteria_data.to_excel(writer, sheet_name='Criteria Analysis', index=False)
            
            # 6. Outliers
            outliers_data = self._create_outliers_dataframe(comparative)
            if outliers_data is not None and not outliers_data.empty:
                outliers_data.to_excel(writer, sheet_name='Outliers', index=False)

            # 7. Advanced Statistics Sheet
            advanced_data = self._create_advanced_statistics_dataframe(comparative)
            if advanced_data is not None and not advanced_data.empty:
                advanced_data.to_excel(writer, sheet_name='Advanced Statistics', index=False)

            # Add after existing sheets (around line 1700)
            if "score_flips" in comparative.get("variants", {}).get(next(iter(comparative.get("variants", {}))), {}):
                # Create score flips summary
                flips_data = []
                for variant_name, variant_data in comparative.get("variants", {}).items():
                    if "score_flips" in variant_data:
                        flips = variant_data["score_flips"]
                        flips_data.append({
                            "Experiment": variant_name,
                            "Fail to Pass": flips.get("fail_to_pass_count", 0),
                            "Pass to Fail": flips.get("pass_to_fail_count", 0),
                            "Major Drops": flips.get("major_drops_count", 0),
                            "Major Gains": flips.get("major_gains_count", 0),
                            "Net Impact": flips.get("net_flip_impact", 0)
                        })
                
                if flips_data:
                    pd.DataFrame(flips_data).to_excel(writer, sheet_name='Score Flips', index=False)


            # Format the Excel file
            workbook = writer.book
            self._format_excel_workbook(workbook)
        
        logger.info(f"Excel report saved to {excel_file}")
    
    def _write_excel_sheets_basic(self, writer, comparative: Dict, recommendations: List):
        """Write Excel sheets using basic pandas functionality"""
        # Summary
        summary_data = self._create_summary_dataframe(comparative)
        if not summary_data.empty:
            summary_data.to_excel(writer, sheet_name='Summary', index=False)
        
        # Distributions
        dist_data = self._create_distribution_dataframe()
        if not dist_data.empty:
            dist_data.to_excel(writer, sheet_name='Score Distributions', index=False)
        
        # Statistical Tests
        stats_data = self._create_statistical_tests_dataframe(comparative)
        if not stats_data.empty:
            stats_data.to_excel(writer, sheet_name='Statistical Tests', index=False)
        
        # Recommendations
        if recommendations:
            rec_data = pd.DataFrame(recommendations)
            # Flatten nested metrics dict
            if 'metrics' in rec_data.columns:
                metrics_df = pd.json_normalize(rec_data['metrics'])
                rec_data = pd.concat([rec_data.drop('metrics', axis=1), metrics_df], axis=1)
            rec_data.to_excel(writer, sheet_name='Recommendations', index=False)
    
    def _create_summary_dataframe(self, comparative: Dict) -> pd.DataFrame:
        """Create summary dataframe for Excel"""
        if not comparative:
            return pd.DataFrame()
            
        rows = []
        
        # Baseline row
        baseline_stats = comparative.get("baseline_stats", {})
        rows.append({
            "Experiment": self.baseline_name + " (BASELINE)",
            "Mean Score": self._safe_get_numeric(baseline_stats.get("mean_score")),
            "Median Score": self._safe_get_numeric(baseline_stats.get("median_score")),
            "Std Dev": self._safe_get_numeric(baseline_stats.get("std_dev")),
            "Has Score 1": "Yes" if baseline_stats.get("has_score_1") else "No",
            "Total Scores": int(self._safe_get_numeric(baseline_stats.get("total_scores")))
        })
        
        # Variant rows
        for variant_name, variant_data in comparative.get("variants", {}).items():
            if "error" not in variant_data:
                dist = variant_data.get("distribution_shift", {})
                cons = variant_data.get("consistency", {})
                
                rows.append({
                    "Experiment": variant_name,
                    "Mean Score": self._safe_get_numeric(baseline_stats.get("mean_score")) + 
                                 self._safe_get_numeric(dist.get("mean_delta")),
                    "Mean Delta": self._safe_get_numeric(dist.get("mean_delta")),
                    "Effect Size": self._safe_get_numeric(dist.get("cohens_d")),
                    "Correlation": self._safe_get_numeric(cons.get("spearman_r")),
                    "P-Value": self._safe_get_numeric(dist.get("mann_whitney_pvalue"), 1.0),
                    "Significant": "Yes" if dist.get("mann_whitney_pvalue", 1) < self.SIGNIFICANCE_LEVEL else "No"
                })
        
        return pd.DataFrame(rows)
    
    def _create_distribution_dataframe(self) -> pd.DataFrame:
        """Create distribution dataframe for Excel"""
        rows = []
        
        for exp_name, exp_data in self.all_results.items():
            if "error" not in exp_data:
                scores = self.extract_scores(exp_data["results"])
                if scores:
                    dist = Counter(scores)
                    
                    rows.append({
                        "Experiment": exp_name,
                        "Is Baseline": "Yes" if exp_name == self.baseline_name else "No",
                        "Count 1": dist.get(1, 0),
                        "Count 2": dist.get(2, 0),
                        "Count 3": dist.get(3, 0),
                        "Count 4": dist.get(4, 0),
                        "Count 5": dist.get(5, 0),
                        "% Score 1": dist.get(1, 0) / len(scores) * 100,
                        "% Score 2": dist.get(2, 0) / len(scores) * 100,
                        "% Score 3": dist.get(3, 0) / len(scores) * 100,
                        "% Score 4": dist.get(4, 0) / len(scores) * 100,
                        "% Score 5": dist.get(5, 0) / len(scores) * 100
                    })
        
        return pd.DataFrame(rows)
    
    def _create_statistical_tests_dataframe(self, comparative: Dict) -> pd.DataFrame:
        """Create statistical tests dataframe for Excel"""
        if not comparative:
            return pd.DataFrame()
            
        rows = []
        
        for variant_name, variant_data in comparative.get("variants", {}).items():
            if "error" not in variant_data:
                dist = variant_data.get("distribution_shift", {})
                cons = variant_data.get("consistency", {})
                trans = variant_data.get("transformation", {})
                
                rows.append({
                    "Experiment": variant_name,
                    "KS Statistic": self._safe_get_numeric(dist.get("ks_statistic"), np.nan),
                    "KS P-Value": self._safe_get_numeric(dist.get("ks_pvalue"), np.nan),
                    "Mann-Whitney U": self._safe_get_numeric(dist.get("mann_whitney_u"), np.nan),
                    "MW P-Value": self._safe_get_numeric(dist.get("mann_whitney_pvalue"), np.nan),
                    "Wilcoxon Stat": self._safe_get_numeric(dist.get("wilcoxon_statistic"), np.nan),
                    "Wilcoxon P-Value": self._safe_get_numeric(dist.get("wilcoxon_pvalue"), np.nan),
                    "Pearson R": self._safe_get_numeric(cons.get("pearson_r"), np.nan),
                    "Spearman R": self._safe_get_numeric(cons.get("spearman_r"), np.nan),
                    "Cohen's Kappa": self._safe_get_numeric(cons.get("cohen_kappa"), np.nan),
                    "Linear R²": self._safe_get_numeric(
                        trans.get("linear_model", {}).get("r_squared"), np.nan
                    ) if isinstance(trans, dict) else np.nan
                })
        
        return pd.DataFrame(rows)
    
    def _create_criteria_analysis_dataframe(self, comparative: Dict) -> Optional[pd.DataFrame]:
        """Create criteria analysis dataframe for Excel"""
        if not comparative:
            return None
            
        rows = []
        
        for variant_name, variant_data in comparative.get("variants", {}).items():
            if "error" not in variant_data:
                criteria = variant_data.get("criteria_analysis", {})
                for criteria_name, criteria_stats in criteria.items():
                    if isinstance(criteria_stats, dict) and "error" not in criteria_stats:
                        rows.append({
                            "Experiment": variant_name,
                            "Criterion": criteria_name,
                            "Baseline Mean": self._safe_get_numeric(criteria_stats.get("baseline_mean")),
                            "Variant Mean": self._safe_get_numeric(criteria_stats.get("variant_mean")),
                            "Delta": self._safe_get_numeric(criteria_stats.get("delta")),
                            "Effect Size": self._safe_get_numeric(criteria_stats.get("effect_size")),
                            "P-Value": self._safe_get_numeric(criteria_stats.get("p_value"), 1.0),
                            "Significant": "Yes" if criteria_stats.get("significant") else "No"
                        })
        
        return pd.DataFrame(rows) if rows else None
    
    def _create_outliers_dataframe(self, comparative: Dict) -> Optional[pd.DataFrame]:
        """Create outliers dataframe for Excel"""
        if not comparative:
            return None
            
        rows = []
        
        for variant_name, variant_data in comparative.get("variants", {}).items():
            if "error" not in variant_data:
                outliers = variant_data.get("top_outliers", [])
                for outlier in outliers:
                    rows.append({
                        "Experiment": variant_name,
                        "ID": outlier.get("id"),
                        "Category": outlier.get("category"),
                        "Baseline Mean": outlier.get("baseline_mean"),
                        "Variant Mean": outlier.get("variant_mean"),
                        "Delta": outlier.get("delta"),
                        "Direction": outlier.get("direction")
                    })
        
        return pd.DataFrame(rows) if rows else None
    
    def _format_excel_workbook(self, workbook):
        """Apply formatting to Excel workbook"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Format each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Format headers
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Apply number formatting
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        if isinstance(cell.value, float):
                            if 'p-value' in str(worksheet.cell(1, cell.column).value).lower():
                                cell.number_format = '0.0000'
                            elif '%' in str(worksheet.cell(1, cell.column).value):
                                cell.number_format = '0.00%'
                            else:
                                cell.number_format = '0.000'
        except Exception as e:
            logger.warning(f"Could not apply Excel formatting: {e}")
    
    def _generate_executive_summary(self, output_dir: str, comparative: Dict, recommendations: List):
        """Generate human-readable executive summary WITH NEW ANALYSES"""
            
        # Add None checks
        if comparative is None:
            comparative = {}
        if recommendations is None:
            recommendations = []

        summary_lines = [
            "# Experiment Analysis Executive Summary",
            f"\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            f"\n**Baseline Experiment**: {self.baseline_name}",
            f"\n**Total Experiments**: {len(self.all_results)}",
        ]
        
        # NEW: Add best variant analysis if available
        if comparative and "best_variant_analysis" in comparative:
            best_analysis = comparative["best_variant_analysis"]
            if best_analysis.get("best_variant"):
                summary_lines.extend([
                    "\n## 🏆 Statistical Winner\n",
                    f"**{best_analysis.get('recommendation', 'No recommendation available')}**\n"
                ])
        
        summary_lines.append("\n## Key Findings\n")
        
        # Problem identification
        baseline_stats = comparative.get("baseline_stats", {}) if comparative else {}
        
        problems = []
        if not baseline_stats.get("has_score_1", False):
            problems.append("- ⚠️ **No Score 1s**: Baseline has no lowest scores (1s)")
        
        mean_score = self._safe_get_numeric(baseline_stats.get("mean_score"))
        if mean_score > 3.5:
            problems.append(f"- ⚠️ **Overscoring**: Mean score is {mean_score:.2f} (>3.5)")
        
        std_dev = self._safe_get_numeric(baseline_stats.get("std_dev"))
        if std_dev < 0.5:
            problems.append(f"- ⚠️ **Low Variance**: Std dev is {std_dev:.2f} (<0.5)")
        
        if problems:
            summary_lines.append("### Problems Identified")
            summary_lines.extend(problems)
        else:
            summary_lines.append("✅ No major problems identified in baseline")
        
        # Recommendations
        if recommendations:
            summary_lines.append("\n## Top Recommendations\n")
            
            for i, rec in enumerate(recommendations[:3], 1):
                summary_lines.append(f"\n### {i}. {rec['experiment']}")
                summary_lines.append(f"- **Priority**: {rec['priority']}")
                summary_lines.append(f"- **Reason**: {rec['reason']}")
                summary_lines.append(f"- **Details**: {rec['details']}")
                
                if rec.get('statistical_significance'):
                    summary_lines.append("- **Statistically Significant**: Yes ✓")

        if comparative:
            summary_lines.append("\n## Enhanced Statistical Analysis\n")
            summary_lines.append(
                "| Experiment | Cohen's d [95% CI] | Cliff's δ | Bayesian P(better) | Power | Practical Sig |"
            )
            summary_lines.append(
                "|------------|-------------------|-----------|-------------------|--------|--------------|"
            )
            
            for variant_name, variant_analysis in comparative.get("variants", {}).items():
                if "error" in variant_analysis:
                    continue
                
                adv = variant_analysis.get("advanced_statistics", {})
                
                # Format Cohen's d with CI
                cohens_ci = adv.get("cohens_d_ci", {})
                cohens_str = f"{cohens_ci.get('estimate', 0):.2f} [{cohens_ci.get('ci_95_lower', 0):.2f}, {cohens_ci.get('ci_95_upper', 0):.2f}]"
                
                # Format Cliff's delta
                cliffs = adv.get("cliffs_delta", {})
                cliffs_str = f"{cliffs.get('value', 0):.2f} ({cliffs.get('interpretation', 'unknown')})"
                
                # Bayesian probability
                bayes = adv.get("bayesian", {})
                bayes_str = f"{bayes.get('prob_variant_better', 0.5):.1%}"
                
                # Power
                power = adv.get("power_analysis", {})
                power_str = f"{power.get('statistical_power', 0):.2f}"
                
                # Practical significance
                prac = adv.get("practical_significance", {})
                prac_str = "Yes" if prac.get("is_practically_significant", False) else "No"
                
                summary_lines.append(
                    f"| {variant_name} | {cohens_str} | {cliffs_str} | {bayes_str} | {power_str} | {prac_str} |"
                )

        # All variants summary table
        if comparative:
            summary_lines.append("\n## Experiment Comparison Matrix\n")
            summary_lines.append("| Experiment | Mean Δ | Correlation | Has 1s | Effect Size | Transformation |")
            summary_lines.append("|------------|--------|------------|--------|-------------|---------------|")
            
            for variant_name, variant_analysis in comparative.get("variants", {}).items():
                if "error" in variant_analysis:
                    summary_lines.append(f"| {variant_name} | ERROR | - | - | - | - |")
                    continue
                    
                dist = variant_analysis.get("distribution_shift", {})
                cons = variant_analysis.get("consistency", {})
                trans = variant_analysis.get("transformation", {})
                
                summary_lines.append(
                    f"| {variant_name} | "
                    f"{self._safe_get_numeric(dist.get('mean_delta')):+.2f} | "
                    f"{self._safe_get_numeric(cons.get('spearman_r')):.3f} | "
                    f"{'Yes' if dist.get('variant_has_1s') else 'No'} | "
                    f"{dist.get('effect_size_interpretation', 'unknown')} | "
                    f"{trans.get('transformation_type', 'unknown') if isinstance(trans, dict) else 'unknown'} |"
                )
            
            # Statistical significance summary
            summary_lines.append("\n## Statistical Significance\n")
            
            sig_count = 0
            for variant_name, variant_analysis in comparative.get("variants", {}).items():
                if "error" in variant_analysis:
                    continue
                    
                dist = variant_analysis.get("distribution_shift", {})
                p_value = self._safe_get_numeric(dist.get("mann_whitney_pvalue"), 1.0)
                if p_value < self.SIGNIFICANCE_LEVEL:
                    sig_count += 1
                    summary_lines.append(
                        f"- **{variant_name}**: Significant difference (p={p_value:.4f})"
                    )
            
            if sig_count == 0:
                summary_lines.append("No experiments showed statistically significant differences")
        
        # Write summary
        summary_file = os.path.join(output_dir, "executive_summary.md")
        with open(summary_file, 'w') as f:
            f.write("\n".join(summary_lines))
        
        logger.info(f"Executive summary written to {summary_file}")
    
    def _generate_visualizations(self, output_dir: str):
        """Generate visualization plots if matplotlib is available"""
        try:
            import matplotlib.pyplot as plt
            import seaborn as sns
            
            sns.set_style("whitegrid")
            plt.rcParams['figure.dpi'] = 100
            
            # Create visualizations
            fig, axes = plt.subplots(2, 2, figsize=(15, 12))
            
            # 1. Score distributions
            self._plot_distributions(axes[0, 0])
            
            # 2. Mean comparison
            self._plot_mean_comparison(axes[0, 1])
            
            # 3. Correlation heatmap
            self._plot_correlation_heatmap(axes[1, 0])
            
            # 4. Effect sizes
            self._plot_effect_sizes(axes[1, 1])
            
            plt.tight_layout()
            plt.savefig(os.path.join(output_dir, "analysis_summary.png"), dpi=150, bbox_inches='tight')
            plt.close()
            
            logger.info("Visualizations saved")
            
        except ImportError as e:
            logger.warning(f"Cannot generate visualizations: {e}")
    
    def _plot_distributions(self, ax):
        """Plot score distributions for all experiments"""
        data = []
        labels = []
        
        for exp_name, exp_data in self.all_results.items():
            if "error" not in exp_data:
                scores = self.extract_scores(exp_data["results"])
                if scores:
                    data.append(scores)
                    label = exp_name
                    if exp_name == self.baseline_name:
                        label += " (baseline)"
                    labels.append(label)
        
        if data:
            parts = ax.violinplot(data, showmeans=True, showmedians=True)
            
            # Color baseline differently
            for i, pc in enumerate(parts['bodies']):
                if i == 0 and self.baseline_name in labels[0]:
                    pc.set_facecolor('#ff7f0e')
                else:
                    pc.set_facecolor('#1f77b4')
            
            ax.set_xticks(range(1, len(labels) + 1))
            ax.set_xticklabels(labels, rotation=45, ha='right')
            ax.set_ylabel('Score')
            ax.set_title('Score Distributions by Experiment')
            ax.grid(True, alpha=0.3)
            ax.set_ylim(0.5, 5.5)
    
    def _plot_mean_comparison(self, ax):
        """Plot mean score comparison with error bars"""
        experiments = []
        means = []
        stds = []
        colors = []
        
        for exp_name, exp_data in self.all_results.items():
            if "error" not in exp_data:
                scores = self.extract_scores(exp_data["results"])
                if scores:
                    experiments.append(exp_name)
                    means.append(np.mean(scores))
                    stds.append(np.std(scores))
                    colors.append('#ff7f0e' if exp_name == self.baseline_name else '#1f77b4')
        
        if experiments:
            x_pos = np.arange(len(experiments))
            ax.bar(x_pos, means, yerr=stds, capsize=5, color=colors, alpha=0.7)
            ax.set_xticks(x_pos)
            ax.set_xticklabels(experiments, rotation=45, ha='right')
            ax.set_ylabel('Mean Score')
            ax.set_title('Mean Scores with Standard Deviation')
            ax.axhline(y=3, color='r', linestyle='--', alpha=0.5, label='Neutral (3)')
            ax.legend()
            ax.grid(True, alpha=0.3)
    
    def _plot_correlation_heatmap(self, ax):
        """Plot correlation heatmap between experiments"""
        # Create correlation matrix
        exp_names = [name for name in self.all_results if "error" not in self.all_results[name]]
        n_exp = len(exp_names)
        
        if n_exp > 1:
            corr_matrix = np.zeros((n_exp, n_exp))
            
            for i, exp1 in enumerate(exp_names):
                for j, exp2 in enumerate(exp_names):
                    if i == j:
                        corr_matrix[i, j] = 1.0
                    else:
                        # Get paired scores between experiments
                        paired = self._get_paired_scores_between(exp1, exp2)
                        if len(paired) > 10:
                            scores1 = [p[0] for p in paired]
                            scores2 = [p[1] for p in paired]
                            corr, _ = self._safe_correlation(scores1, scores2, 'spearman')
                            corr_matrix[i, j] = corr if not np.isnan(corr) else 0
            
            im = ax.imshow(corr_matrix, cmap='coolwarm', vmin=-1, vmax=1)
            ax.set_xticks(np.arange(n_exp))
            ax.set_yticks(np.arange(n_exp))
            ax.set_xticklabels(exp_names, rotation=45, ha='right')
            ax.set_yticklabels(exp_names)
            ax.set_title('Experiment Correlation Matrix (Spearman)')
            
            # Add text annotations
            for i in range(n_exp):
                for j in range(n_exp):
                    text = ax.text(j, i, f'{corr_matrix[i, j]:.2f}',
                                 ha="center", va="center", color="black", fontsize=8)
     
            plt.colorbar(im, ax=ax)
    
    def _get_paired_scores_between(self, exp1: str, exp2: str) -> List[Tuple[int, int]]:
        """Get paired scores between any two experiments"""
        paired_scores = []
        
        if exp1 not in self.all_results or exp2 not in self.all_results:
            return paired_scores
        
        results1 = self.all_results[exp1]["results"]
        results2 = self.all_results[exp2]["results"]
        
        # Create ID mapping
        results1_by_id = {r["id"]: r for r in results1}
        results2_by_id = {r["id"]: r for r in results2}
        
        # Find common IDs
        common_ids = set(results1_by_id.keys()) & set(results2_by_id.keys())
        
        for id_ in common_ids:
            r1 = results1_by_id[id_]
            r2 = results2_by_id[id_]
            
            for cs_key in ["case_study_1_llm", "case_study_2_llm"]:
                scores1 = self._convert_scores(r1.get(cs_key, []))
                scores2 = self._convert_scores(r2.get(cs_key, []))
                
                if len(scores1) == len(scores2):
                    for s1, s2 in zip(scores1, scores2):
                        paired_scores.append((s1, s2))
        
        return paired_scores
    
    def _plot_effect_sizes(self, ax):
        """Plot effect sizes for each variant"""
        variants = []
        effect_sizes = []
        colors = []
        
        for variant_name in self.all_results:
            if variant_name != self.baseline_name and "error" not in self.all_results[variant_name]:
                dist = self.calculate_distribution_shift(variant_name)
                if "error" not in dist:
                    variants.append(variant_name)
                    cohens_d = self._safe_get_numeric(dist.get("cohens_d"))
                    effect_sizes.append(cohens_d)
                    
                    # Color based on effect size
                    abs_d = abs(cohens_d)
                    if abs_d < 0.2:
                        colors.append('gray')
                    elif abs_d < 0.5:
                        colors.append('yellow')
                    elif abs_d < 0.8:
                        colors.append('orange')
                    else:
                        colors.append('red')
        
        if variants:
            y_pos = np.arange(len(variants))
            ax.barh(y_pos, effect_sizes, color=colors, alpha=0.7)
            ax.set_yticks(y_pos)
            ax.set_yticklabels(variants)
            ax.set_xlabel("Cohen's d")
            ax.set_title("Effect Sizes vs Baseline")
            
            # Add reference lines
            ax.axvline(x=-0.2, color='gray', linestyle='--', alpha=0.5)
            ax.axvline(x=0.2, color='gray', linestyle='--', alpha=0.5)
            ax.axvline(x=-0.5, color='gray', linestyle='--', alpha=0.5)
            ax.axvline(x=0.5, color='gray', linestyle='--', alpha=0.5)
            ax.axvline(x=-0.8, color='gray', linestyle='--', alpha=0.5)
            ax.axvline(x=0.8, color='gray', linestyle='--', alpha=0.5)
            ax.axvline(x=0, color='black', linestyle='-', alpha=0.8)
            
            ax.grid(True, alpha=0.3)